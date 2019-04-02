#!/usr/bin/env python
#--*-- coding: utf-8 --*--
from flask import Flask, render_template, flash, redirect, url_for, session, request, logging, Blueprint, g, send_file
from wtforms import Form, StringField, TextAreaField, PasswordField, validators
from passlib.hash import sha256_crypt
from functools import wraps
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.datastructures import ImmutableMultiDict
from werkzeug.utils import secure_filename
from .db_connection import getDB
db = getDB()
from .auth import is_logged_in
import json
import traceback
from flask_mail import Mail, Message
from . import generic_functions
GF=generic_functions.GenericFunctions()
from flask import current_app as app
#app=Flask(__name__)
import sys
import time
import os
import random
import app_config as cfg
import datetime
import time
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Color, colors, PatternFill, Border, Alignment
from openpyxl.cell import Cell
import subprocess

bp = Blueprint('task', __name__, url_prefix='/task')

@bp.route('/getSupervisor',methods=['GET','POST'])
@is_logged_in
def getSupervisor():
    response={}
    try:
        if request.method=='POST':
            flag,data=GF.toDict(request.form,'post')
            if flag:
                condition=""
                #if int(data['user_type_id'])==2:
                #    condition=" and user_id=%s"%data['user_id']
                if int(data['user_type_id'])==6: #si es admin/sup se incluye el usuario en lista de supervisores
                    condition=" and user_type_id in (2,3,6)"
                else:
                    condition=" and user_type_id in (2,3)"
                supervisor=db.query("""
                    select
                        user_id as supervisor_id,
                        name
                    from
                        system.user
                    where
                        --user_type_id in (2,6) and
                    company_id=%s
                    and enabled in (1,3) %s
                    order by name
                """%(int(data['company_id']),condition)).dictresult()
                response['data']=supervisor
                response['success']=True
            else:
                response['success']=False
                response['msg_response']='Ocurrió un error al intentar obtener la información.'
        else:
            response['success']=False
            response['msg_response']='Inténtelo de nuevo.'
    except:
        response['success']=False
        response['msg_response']='Ocurrió un error.'
        exc_info = sys.exc_info()
        app.logger.info(traceback.format_exc(exc_info))
    return json.dumps(response)

@bp.route('/getAssignee',methods=['GET','POST'])
@is_logged_in
def getAssignee():
    response={}
    try:
        # app.logger.info('Info')
        if request.method=='POST':
            flag,data=GF.toDict(request.form,'post')
            if flag:
                condition=""
                user_type=""
                #if int(data['user_type_id'])==3:
                #    condition=" and user_id=%s"%data['user_id']
                # if int(data['user_type_id'])==6:
                #     user_type=" user_type_id in (2,3) "
                # else:
                #     user_type=" user_type_id = 3 "
                if int(data['user_type_id'])==6:
                    user_type=" user_type_id in (2,3,6) "
                else:
                    user_type=" user_type_id in (2,3) "
                assignee=db.query("""
                    select
                        user_id as assignee_id,
                        name
                    from
                        system.user
                    where
                        --user_type_id=3
                        %s
                    and company_id=%s %s
                    and enabled in (1,3)
                    order by name
                """%(user_type,int(data['company_id']),condition)).dictresult()
                response['data']=assignee
                response['success']=True
            else:
                response['success']=False
                response['msg_response']='Ocurrió un error al intentar obtener la información.'
        else:
            response['success']=False
            response['msg_response']='Inténtelo de nuevo.'
    except:
        response['success']=False
        response['msg_response']='Ocurrió un error.'
        exc_info = sys.exc_info()
        app.logger.info(traceback.format_exc(exc_info))
    return json.dumps(response)

@bp.route('/saveTask',methods=['GET','POST'])
@is_logged_in
def saveTask():
    response={}
    try:
        flag,data=GF.toDict(request.form,'post')
        if flag:

            valid=True
            for k,v in data.iteritems():
                if v=="" or v==None:
                    if k!='description':
                        valid=False
            if valid:
                deadline=time.strptime(data['deadline'],"%Y-%m-%d")
                supervisor_deadline=time.strptime(data['supervisor_deadline'],"%Y-%m-%d")
                assignee_deadline=time.strptime(data['assignee_deadline'],"%Y-%m-%d")
                if assignee_deadline<=supervisor_deadline and supervisor_deadline<=deadline:
                    data['status_id']=1
                    data['created']='now'
                    data['last_updated']='now'
                    data['user_last_updated']=data['user_id']
                    data['created_by']=data['user_id']
                    data['supervisor_deadline']="%s 23:59:59"%data['supervisor_deadline']
                    data['assignee_deadline']="%s 23:59:59"%data['assignee_deadline']
                    data['deadline']="%s 23:59:59"%data['deadline']

                    new_task=db.insert('task.task',data)
                    app.logger.info("Se crea tarea %s, company_id:%s"%(data['name'],data['company_id']))
                    documents=json.loads(data['document'])
                    for x in documents:
                        doc={
                            'task_id':new_task['task_id'],
                            'name':x['name'],
                            'document_type_id':x['document_type_id'],
                            'description':x['description']
                        }
                        db.insert('task.document',doc)

                    task_info=db.query("""
                        select
                            (select name from system.user where user_id=a.supervisor_id) as supervisor,
                            a.name,
                            (select name from system.user where user_id=a.assignee_id) as assignee,
                            a.description,
                            to_char(a.assignee_deadline,'DD-MM-YYYY HH24:MI:SS') as assignee_deadline,
                            to_char(a.supervisor_deadline,'DD-MM-YYYY HH24:MI:SS') as supervisor_deadline,
                            to_char(a.deadline, 'DD-MM-YYYY HH24:MI:SS') as deadline
                        from
                            task.task a
                        where a.task_id=%s
                    """%new_task['task_id']).dictresult()[0]

                    company_name=db.query("""
                        select name from system.company where company_id=%s
                    """%data['company_id']).dictresult()[0]
                    task_info['company']=company_name['name']
                    message=db.query("""
                        select * from template.generic_template where type_id=24
                    """).dictresult()[0]
                    task_info['link']=cfg.host
                    task_info['mail_img']=cfg.mail_img
                    msg=message['body'].format(**task_info)

                    if data['notify_admin']==True:
                        check_sup_type=db.query("""
                            select user_type_id from system.user where user_id=%s
                        """%data['supervisor_id']).dictresult()[0]
                        if check_sup_type['user_type_id']==2:
                            recipients=db.query("""
                                select email from system.user
                                where (company_id=%s and user_type_id in (1,6)) or (user_id in (%s,%s))
                            """%(data['company_id'],data['assignee_id'],data['supervisor_id'])).dictresult()
                        else:
                            recipients=db.query("""
                                select email from system.user where user_id in (%s,%s)
                            """%(data['assignee_id'],data['supervisor_id'])).dictresult()
                    else:
                        recipients=db.query("""
                            select email from system.user where user_id in (%s,%s)
                        """%(data['assignee_id'],data['supervisor_id'])).dictresult()
                    mail_recipients=[]
                    for r in recipients:
                        mail_recipients.append(r['email'])
                    GF.sendMail(message['subject'].format(**task_info),msg,mail_recipients)

                    # #assignee
                    # message=db.query("""
                    #     select * from template.generic_template where type_id=1
                    # """).dictresult()[0]
                    # recipient=db.query("""
                    #     select email from system.user where user_id=%s
                    # """%data['assignee_id']).dictresult()[0]['email']
                    # task_info['link']=cfg.host
                    # task_info['mail_img']=cfg.mail_img
                    # msg=message['body'].format(**task_info)
                    # GF.sendMail(message['subject'].format(**task_info),msg,recipient)
                    #
                    # #supervisor
                    # message_sup=db.query("""
                    #     select * from template.generic_template where type_id=18
                    # """).dictresult()[0]
                    # recipient_sup=db.query("""
                    #     select email from system.user where user_id=%s
                    # """%data['supervisor_id']).dictresult()[0]['email']
                    # msg_sup=message_sup['body'].format(**task_info)
                    # GF.sendMail(message_sup['subject'].format(**task_info),msg_sup,recipient_sup)
                    #
                    # supervisor_type=db.query("""
                    #     select user_type_id from system.user where user_id=%s
                    # """%data['supervisor_id']).dictresult()[0]['user_type_id']
                    # if supervisor_type==2: #si es solo supervisor, si es supervisor/admin, no es necesario enviar correo a administrador
                    #     message_admin=db.query("""
                    #         select * from template.generic_template where type_id=24
                    #     """).dictresult()[0]
                    #     recipient_admin=db.query("""
                    #         select name, email from system.user where company_id=%s and user_type_id in (1,6)
                    #     """%data['company_id']).dictresult()[0]
                    #     task_info['admin']=recipient_admin['name']
                    #     task_info['mail_img']=cfg.mail_img
                    #     msg_admin=message_admin['body'].format(**task_info)
                    #     GF.sendMail(message_admin['subject'].format(**task_info),msg_admin,recipient_admin['email'])

                    response['success']=True
                    response['msg_response']='La tarea ha sido creada.'
                else:
                    response['success']=False
                    msg=""
                    if assignee_deadline>supervisor_deadline:
                        msg+="La fecha límite del auxiliar debe ser igual o menor a la fecha límite del supervisor.<br>"
                    if supervisor_deadline>deadline:
                        msg+="La fecha límite del supervisor debe ser menor o igual a la fecha de vencimiento de la tarea."
                    response['msg_response']=msg
            else:
                response['success']=False
                response['msg_response']="Existen campos vacíos o incompletos."
        else:
            response['success']=False
            response['msg_response']='Ocurrió un error al procesar los datos, inténtelo de nuevo.'
    except:
        response['success']=False
        response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo.'
        exc_info=sys.exc_info()
        app.logger.info(traceback.format_exc(exc_info))
    return json.dumps(response)

@bp.route("/getTask/<option>", methods=['GET','POST'])
@is_logged_in
def getTask(option):
    response={}
    try:
        if request.method=='POST':
            if option=='all':
                user_type_id=int(request.form['user_type_id'])
                user_id=int(request.form['user_id'])
                show_hidden_tasks=request.form['show_hidden_tasks']
                user=""
                deadline=""
                filter=json.loads(request.form['filter'])

                first=request.form['first']
                filters=""
                now=datetime.datetime.now()

                if show_hidden_tasks=='false':
                    hidden_tasks=" and a.hidden=%s"%show_hidden_tasks
                else:
                    hidden_tasks=""

                if first=='true':
                    if user_type_id==2:
                        filter['supervisor_id']=user_id
                    elif user_type_id==3:
                        filter['assignee_id']=user_id
                for key,value in filter.iteritems():
                    if value!=-1:

                        if key[-3:]=='_id':
                            if user_type_id!=2:
                                filters+=" and a.%s = %s"%(key,value)
                            else:
                                if key!='supervisor_id':
                                    filters+=" and a.%s = %s"%(key,value)
                        elif key=='to' or key=='from':
                            if value=="":
                                if key=="to" and filter['from']!="":
                                    filter['to']=filter['from']
                                elif key=="to" and filter['from']=="":
                                    filter['from']="%s-%s-01"%(now.year,str(now.month).zfill(2))
                                    filter['to']="%s-%s-%s"%(now.year,str(now.month).zfill(2),now.day)
                                elif key=="from" and filter['to']!="":
                                    filter['from']=filter['to']
                                elif key=="from" and filter['to']=="":
                                    filter['from']="%s-%s-01"%(now.year,str(now.month).zfill(2))
                                    filter['to']="%s-%s-%s"%(now.year,str(now.month).zfill(2),now.day)
                        elif key=='search' and value!="":
                            filters+=" and a.name||a.description ilike '%%%s%%'"%value
                    else:
                        if key=='from':
                            filter['from']="%s-%s-01"%(now.year,str(now.month).zfill(2))
                            filter['to']="%s-%s-%s"%(now.year,str(now.month).zfill(2),now.day)

                if 'from' not in filter:
                    filter['from']="%s-%s-01"%(now.year,str(now.month).zfill(2))
                if 'to' not in filter:
                    filter['to']="%s-%s-%s"%(now.year,str(now.month).zfill(2),now.day)

                if filter['from']!="" and filter['to']!="":
                    cfrom=time.strptime(filter['from'],"%Y-%m-%d")
                    cto=time.strptime(filter['to'],"%Y-%m-%d")
                    if cfrom>cto:
                        filter['from']=filter['to']

                if filter['date_type']==1:
                    filters+=" and a.created between '%s 00:00:00' and '%s 23:59:59'"%(filter['from'],filter['to'])

                response['first']=False #indicates there are no tasks before the current month
                response['last']=False #indicates there are no tasks after the current month
                if user_type_id in (1,4,5,6):
                    user=""
                    deadline="to_char(a.deadline,'DD-MM-YYYY') as deadline"
                    if filter['date_type']==2:
                        if first=='true':
                            older_task=db.query("""
                                select to_char(deadline,'YYYY-MM-DD') as deadline from task.task
                                where company_id=%s and status_id in (1,6) order by deadline asc limit 1
                            """%int(request.form['company_id'])).dictresult()

                            last_task=db.query("""
                                select to_char(deadline,'YYYY-MM-DD') as deadline from task.task
                                where company_id=%s and status_id in (1,6) order by deadline desc limit 1
                            """%int(request.form['company_id'])).dictresult()

                            date_filter=""
                            if older_task!=[]:
                                date_filter=" and a.deadline between '%s 00:00:00'"%older_task[0]['deadline']
                                # filters+=" and a.deadline between '%s 00:00:00' and '%s 23:59:59' "%(older_task[0]['deadline'],filter['to'])
                                response['first']=True #indicates there are tasks before the current month
                                response['older_task']=older_task[0]['deadline']
                            else:
                                date_filter=" and a.deadline between '%s 00:00:00'"%filter['from']
                                # filters+=" and a.deadline between '%s 00:00:00' and '%s 23:59:59'"%(filter['from'],filter['to'])
                            if last_task!=[]:
                                date_filter+=" and '%s 23:59:59'"%last_task[0]['deadline']
                                response['last']=True #indicates there are tasks after the current month
                                response['last_task']=last_task[0]['deadline']
                            else:
                                date_filter+=" and '%s 23:59:59'"%filter['to']
                            filters+=date_filter

                        else:
                            filters+=" and a.deadline between '%s 00:00:00' and '%s 23:59:59'"%(filter['from'],filter['to'])
                elif user_type_id==2:
                    user=" and (supervisor_id=%s or assignee_id=%s)"%(user_id,user_id)
                    #user= "and supervisor_id=%s"%user_id
                    deadline="to_char(a.supervisor_deadline,'DD-MM-YYYY') as deadline"
                    if filter['date_type']==2:
                        if first=='true':
                            older_task=db.query("""
                                select to_char(supervisor_deadline,'YYYY-MM-DD') as supervisor_deadline from task.task
                                where supervisor_id=%s and company_id=%s and status_id in (1,6) order by supervisor_deadline asc limit 1
                            """%(int(request.form['user_id']),int(request.form['company_id']))).dictresult()
                            last_task=db.query("""
                                select to_char(supervisor_deadline,'YYYY-MM-DD') as supervisor_deadline from task.task
                                where supervisor_id=%s and company_id=%s and status_id in (1,6) order by supervisor_deadline desc limit 1
                            """%(int(request.form['user_id']),int(request.form['company_id']))).dictresult()
                            date_filter=""
                            if older_task!=[]:
                                date_filter=" and a.supervisor_deadline between '%s 00:00:00'"%older_task[0]['supervisor_deadline']
                                # filters+=" and a.supervisor_deadline between '%s 00:00:00' and '%s 23:59:59'"%(older_task[0]['supervisor_deadline'],filter['to'])
                                response['first']=True
                                response['older_task']=older_task[0]['supervisor_deadline']
                            else:
                                date_filter=" and a.supervisor_deadline between '%s 00:00:00'"%filter['from']
                                # filters+=" and a.supervisor_deadline between '%s 00:00:00' and '%s 23:59:59'"%(filter['from'],filter['to'])
                            if last_task!=[]:
                                date_filter+=" and '%s 23:59:59'"%last_task[0]['supervisor_deadline']
                                response['last']=True
                                response['last_task']=last_task[0]['supervisor_deadline']
                            else:
                                date_filter+=" and '%s 23:59:59'"%filter['to']
                            filters+=date_filter
                        else:
                            filters+=" and a.supervisor_deadline between '%s 00:00:00' and '%s 23:59:59'"%(filter['from'],filter['to'])

                elif user_type_id==3:
                    user=" and assignee_id=%s"%user_id
                    deadline="to_char(a.assignee_deadline,'DD-MM-YYYY') as deadline"
                    if filter['date_type']==2:
                        if first=='true':
                            older_task=db.query("""
                                select to_char(assignee_deadline,'YYYY-MM-DD') as assignee_deadline from task.task
                                where assignee_id=%s and company_id=%s and status_id in (1,6) order by assignee_deadline asc limit 1
                            """%(int(request.form['user_id']),int(request.form['company_id']))).dictresult()
                            last_task=db.query("""
                                select to_char(assignee_deadline,'YYYY-MM-DD') as assignee_deadline from task.task
                                where assignee_id=%s and company_id=%s and status_id in (1,6) order by assignee_deadline desc limit 1
                            """%(int(request.form['user_id']),int(request.form['company_id']))).dictresult()
                            date_filter=""
                            if older_task!=[]:
                                date_filter=" and a.assignee_deadline between '%s 00:00:00'"%older_task[0]['assignee_deadline']
                                # filters+=" and a.assignee_deadline between '%s 00:00:00' and '%s 23:59:59'"%(older_task[0]['assignee_deadline'],filter['to'])
                                response['first']=True
                                response['older_task']=older_task[0]['assignee_deadline']
                            else:
                                date_filter=" and a.assignee_deadline between '%s 00:00:00'"%filter['from']
                                # filters+=" and a.assignee_deadline between '%s 00:00:00' and '%s 23:59:59'"%(filter['from'],filter['to'])
                            if last_task!=[]:
                                date_filter+=" and '%s 23:59:59'"%last_task[0]['assignee_deadline']
                                response['last']=True
                                response['last_task']=last_task[0]['assignee_deadline']
                            else:
                                date_filter+=" and '%s 23:59:59'"%filter['to']
                            filters+=date_filter
                        else:
                            filters+=" and a.assignee_deadline between '%s 00:00:00' and '%s 23:59:59'"%(filter['from'],filter['to'])

                if filter['date_type']==1:
                    order_by="a.created"
                else:
                    if user_type_id in (1,4,5,6):
                        order_by="a.deadline"
                    elif user_type_id == 2:
                        order_by="a.supervisor_deadline"
                    else:
                        order_by="a.assignee_deadline"

                user=""

                task=db.query("""
                    select
                        a.task_id,
                        a.name,
                        a.description,
                        %s,
                        a.supervisor_id,
                        a.assignee_id,
                        to_char(a.created,'DD-MM-YYYY') as created,
                        a.company_id,
                        a.status_id,
                        (select name from system.user where user_id=a.assignee_id) as assignee,
                        (select name from system.user where user_id=a.supervisor_id) as supervisor,
                        b.description as status,
                        a.res_dec_by
                    from
                        task.status b,
                        task.task a
                    where
                        a.company_id=%s
                    and a.status_id=b.status_id
                    %s %s %s
                    order by %s asc
                    offset %s limit %s
                """%(deadline,int(request.form['company_id']),user,filters,hidden_tasks,order_by,int(request.form['start']),int(request.form['length']))).dictresult()

                total=db.query("""
                    select
                        count(*)
                    from
                        task.status b,
                        task.task a
                    where
                        a.company_id=%s
                    and a.status_id=b.status_id
                    %s %s %s
                """%(int(request.form['company_id']),user,filters,hidden_tasks)).dictresult()

                response['data']=task
                response['recordsTotal']=total[0]['count']
                response['recordsFiltered']=total[0]['count']
                response['success']=True
            else:
                user_type_id=int(request.form['user_type_id'])
                deadline=""
                if user_type_id in (1,4,5,6):
                    deadline="to_char(a.deadline,'DD-MM-YYYY') as deadline"
                elif user_type_id==2:
                    deadline="to_char(a.supervisor_deadline,'DD-MM-YYYY') as deadline"
                elif user_type_id==3:
                    deadline="to_char(a.assignee_deadline,'DD-MM-YYYY') as deadline"
                task_id=int(option,16)/cfg.factor_tt
                task=db.query("""
                    select
                        a.task_id,
                        a.name,
                        a.description,
                        %s,
                        a.supervisor_id,
                        a.assignee_id,
                        to_char(a.created,'DD-MM-YYYY') as created,
                        a.company_id,
                        a.status_id,
                        (select name from system.user where user_id=a.assignee_id) as assignee,
                        (select name from system.user where user_id=a.supervisor_id) as supervisor,
                        b.description as status,
                        a.res_dec_by
                    from
                        task.status b,
                        task.task a
                    where
                        a.task_id=%s
                        and a.status_id=b.status_id
                """%(deadline,int(task_id))).dictresult()
                response['data']=task
                response['recordsTotal']=1
                response['recordsFiltered']=1
                response['success']=True

        else:
            response['success']=False
            response['msg_response']='Cargue la página nuevamente.'
    except:
        response['success']=False
        response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo.'
        exc_info=sys.exc_info()
        app.logger.info(traceback.format_exc(exc_info))
    return json.dumps(response)

@bp.route('/getTaskDetails',methods=['GET','POST'])
@is_logged_in
def getTaskDetails():
    response={}
    try:
        flag,data=GF.toDict(request.form,'post')

        if flag:
            deadline=""
            if data['user_type_id'] in (1,4,5,6):
                deadline="a.deadline"
            elif data['user_type_id']==2:
                deadline="a.supervisor_deadline"
            elif data['user_type_id']==3:
                deadline="a.assignee_deadline"
            task=db.query("""
                select
                    a.task_id,
                    a.name,
                    a.description,
                    to_char(%s,'DD-MM-YYYY') as deadline,
                    (select name from system.user where user_id=a.assignee_id) as assignee,
                    (select name from system.user where user_id=a.supervisor_id) as supervisor,
                    to_char(a.created,'DD-MM-YYYY') as created,
                    b.description as status,
                    to_char(a.resolved_date,'DD-MM-YYYY HH24:MI:SS') as resolved_date,
                    (select name from system.user where user_id=a.created_by) as created_by,
                    a.comments,
                    a.status_id
                from
                    task.status b,
                    task.task a
                where
                    a.task_id=%s
                and a.status_id=b.status_id
                and a.company_id=%s
            """%(deadline,data['task_id'],data['company_id'])).dictresult()[0]

            documents=db.query("""
                select
                    a.document_id,
                    a.name,
                    b.document_type,
                    a.file_path,
                    to_char(a.loaded,'DD-MM-YYYY HH24:MI:SS') as loaded,
                    a.file_name
                from
                    task.document_type b,
                    task.document a
                where
                    a.task_id=%s
                and a.document_type_id=b.document_type_id
            """%data['task_id']).dictresult()

            if data['from']=='details' or data['from']=='decline':
                doc_list=""
                if documents!=[]:
                    for x in documents:
                        if x['file_path']=="":
                            doc_list+="<li>%s (%s)</li>"%(x['name'],x['document_type'])
                        else:
                            # doc_list+="<li>%s (%s) <br> cargado %s</li>"%(x['name'],x['document_type'],x['loaded'])
                            if task['status_id']==4:
                                random_number=int(random.random()*100000)
                                fname,ext=os.path.splitext(x['file_name'])

                                doc_list+="""
                                    <div style="display:inline-block"><li>%s (%s) <br> cargado %s</li><a href="/task/downloadEvidence/%s_%s%s" target="_blank" role="button" class="btn btn-danger detail-ev-buttons" data-toggle="tooltip" title="Descargar %s"><i class="fa fa-file-text-o"></i></a></div>
                                """%(x['name'],x['document_type'],x['loaded'],random_number,x['document_id'],ext,x['name'])
                            else:
                                doc_list+="<li>%s (%s) <br> cargado %s</li>"%(x['name'],x['document_type'],x['loaded'])

                if task['resolved_date']=='01-01-1900 00:00:00':
                    resolved_date="--"
                else:
                    resolved_date=task['resolved_date']
                str_other_dates=""
                if data['user_type_id']!=3:
                    other_dates=db.query("""
                        select
                            to_char(assignee_deadline,'DD-MM-YYYY') as assignee_deadline,
                            to_char(supervisor_deadline, 'DD-MM-YYYY') as supervisor_deadline
                        from
                            task.task
                        where
                            task_id=%s
                    """%data['task_id']).dictresult()[0]
                    if data['user_type_id'] in (1,4,5,6):
                        str_other_dates="<b>Fecha límite auxiliar:</b> %s<br><b>Fecha límite supervisor:</b> %s<br>"%(other_dates['assignee_deadline'],other_dates['supervisor_deadline'])
                    else:
                        str_other_dates="<b>Fecha límite auxiliar:</b> %s<br>"%other_dates['assignee_deadline']

                html="""
                    <p><b>Nombre:</b> %s <br> <b>Descripción:</b> %s <br> %s <b>Fecha límite:</b> %s <br> <b>Supervisa:</b> %s <br> <b>Asignado a:</b> %s <br> <b>Creada:</b> %s <br> <b>Creada por:</b> %s <br> <b>Status:</b> %s <br> <b>Fecha en que se resolvió:</b> %s <br> <b>Evidencias necesarias:</b> <ul>%s</ul></p>
                """%(task['name'],task['description'],str_other_dates,task['deadline'],task['supervisor'],task['assignee'],task['created'],task['created_by'],task['status'],resolved_date,doc_list)
            elif data['from']=='resolve':
                html="""
                    <p><b>Nombre:</b> %s <br> <b>Descripción:</b> %s <br> <b>Fecha límite:</b> %s <br> <b>Supervisa:</b> %s <br> <b>Asignado a:</b> %s <br> <b>Status:</b> %s </p>
                """%(task['name'],task['description'],task['deadline'],task['supervisor'],task['assignee'],task['status'])

                documents=db.query("""
                    select
                        a.document_id,
                        a.name,
                        b.document_type,
                        a.description,
                        b.document_extension,
                        a.file_path,
                        to_char(a.loaded,'DD-MM-YYYY HH24:MI:SS') as loaded,
                        a.size
                    from
                        task.document_type b,
                        task.document a
                    where
                        a.task_id=%s
                    and a.document_type_id=b.document_type_id
                """%data['task_id']).dictresult()
                html_docs=[]
                if documents!=[]:
                    for d in documents:
                        doc_ext=eval(d['document_extension'])
                        str_doc_ext=','.join(e for e in doc_ext)
                        data_size=''
                        loaded_date=''
                        classes=''
                        if task['status_id']==6:
                            if d['loaded']!='01-01-1900 00:00:00':
                                loaded_date='cargado %s'%d['loaded']
                                classes='valid-file-field'
                                data_size='data-size="%s"'%d['size']
                            else:
                                classes='file-input'
                        else:
                            classes='file-input'
                            data_size='data-size="0"'
                        document_type_name=d['document_type']
                        d['document_type']=d['document_type'].replace(" ","_")

                        input="""<div><label for="input%s%s" class="file-input-label">%s (%s) %s</label><input type="file" id="input%s%s" name="file_%s" lang="es" pattern="%s" class="file-evidence %s" data-toggle="tooltip" title="%s" %s><span id="spninput%s%s" class="error-msg">Error</span><div>"""%(d['document_type'],d['document_id'],d['name'],document_type_name,loaded_date,d['document_type'],d['document_id'],d['document_id'],str_doc_ext,classes,d['description'],data_size,d['document_type'],d['document_id'])
                        html_docs.append(input)
                response['html_docs']=html_docs
                response['comments']=task['comments']

            elif data['from']=='check':
                assignee_info=db.query("""
                    select
                        to_char(assignee_deadline,'DD-MM-YYYY HH24:MI:SS') as date,
                        comments
                    from task.task
                    where task_id=%s
                """%data['task_id']).dictresult()[0]
                html="""
                    <p><b>Nombre:</b> %s <br> <b>Descripción:</b> %s <br> <b>Fecha límite:</b> %s <br> <b>Supervisa:</b> %s <br> <b>Asignado a:</b> %s <br> <b>Fecha límite de auxiliar:</b> %s <br> <b>Fecha en que se resolvió:</b> %s <br> <b>Comentarios auxiliar:</b> %s </p>
                """%(task['name'],task['description'],task['deadline'],task['supervisor'],task['assignee'],assignee_info['date'],task['resolved_date'],assignee_info['comments'])

                documents=db.query("""
                    select name, document_id,file_name
                    from task.document
                    where task_id=%s
                """%data['task_id']).dictresult()
                buttons=""
                random_number=int(random.random()*100000)
                for d in documents:
                    fname,ext=os.path.splitext(d['file_name'])
                    buttons+="""
                        <a href="/task/downloadEvidence/%s_%s%s" target="_blank" role="button" class="btn btn-success" data-toggle="tooltip" title="Descargar %s">%s</a>
                    """%(random_number,d['document_id'],ext,d['name'],d['name'])
                response['evidence']=buttons

            elif data['from']=='check_declined':
                declined_by=db.query("""
                    select
                        a.res_dec_by,
                        b.user_type_id
                    from
                        system.user b,
                        task.task a
                    where
                        a.task_id=%s
                    and a.res_dec_by=b.user_id
                """%data['task_id']).dictresult()[0]

                assignee_info=db.query("""
                    select
                        to_char(assignee_deadline,'DD-MM-YYYY HH24:MI:SS') as date,
                        to_char(last_updated,'DD-MM-YYYY HH24:MI:SS') as last_updated,
                        declining_cause
                    from task.task
                    where task_id=%s
                """%data['task_id']).dictresult()[0]
                if declined_by['user_type_id']==3 and data['user_type_id']==2: #cuando la tarea fue declinada por auxiliar y va a revisar supervisor, mostrará ventana donde solo permite cambiar auxiliar y/o descripción de la tarea
                    response['declined_by']='assignee'
                    html="""
                        <p><b>Nombre:</b> %s <br> <b>Descripción:</b> %s <br> <b>Fecha límite:</b> %s <br> <b>Supervisa:</b> %s <br> <b>Asignado a:</b> %s <br> <b>Fecha límite de auxiliar:</b> %s <br> <b>Fecha en que se declinó:</b> %s <br> <b>Comentarios auxiliar:</b> %s </p>
                    """%(task['name'],task['description'],task['deadline'],task['supervisor'],task['assignee'],assignee_info['date'],assignee_info['last_updated'],assignee_info['declining_cause'])
                else:
                    #en los demás casos
                    response['declined_by']='supervisor'

                    if data['user_id']==declined_by['res_dec_by'] or data['user_type_id']==2:
                        # si el usuario que declina es el mismo que desea revisar no lo permite
                        #si fue declinado por un supervisor, no puede ser revisado por otro supervisor
                        response['allow_check']=False
                        html=""
                    else:
                        response['allow_check']=True
                        html="""<p><b>Nombre:</b> %s <br> <b>Descripción:</b> %s"""%(task['name'],task['description'])
                        deadlines=db.query("""
                            select
                                to_char(assignee_deadline,'YYYY-MM-DD') as assignee_deadline,
                                to_char(supervisor_deadline,'YYYY-MM-DD') as supervisor_deadline,
                                to_char(deadline,'YYYY-MM-DD') as deadline
                            from
                                task.task
                            where
                                task_id=%s
                        """%data['task_id']).dictresult()[0]
                        response['deadlines']=deadlines

            response['data']=html
            response['success']=True
        else:
            response['success']=False
            response['msg_response']="Ocurrió un error al intentar obtener los detalles de la tarea, favor de intentarlo de nuevo."
    except:
        response['success']=False
        response['msg_response']="Ocurrió un error, favor de intentarlo de nuevo."
        exc_info=sys.exc_info()
        app.logger.info(traceback.format_exc(exc_info))
    return json.dumps(response)

@bp.route('/getDocumentType', methods=['GET','POST'])
@is_logged_in
def getDocumentType():
    response={}
    try:
        doc_type=db.query("""
            select
                document_type_id,
                document_type
            from
                task.document_type
        """).dictresult()
        response['data']=doc_type
        response['success']=True
    except:
        response['success']=False
        response['msg_response']="Ocurrió un error, favor de intentarlo de nuevo."
        exc_info=sys.exc_info()
        app.logger.info(traceback.format_exc(exc_info))
    return json.dumps(response)

UPLOAD_FOLDER = '%s'%cfg.task_path
ALLOWED_EXTENSIONS = set(['txt', 'pdf'])

def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@bp.route('/resolveTask', methods=['GET','POST'])
@is_logged_in
def resolveTask():
    response={}
    try:
        data=request.form.to_dict()
        files_list=eval(data['files_list'])
        frm=request.files
        ev_cont=1
        folder=db.query("""
            select task_folder from system.company where company_id=%s
        """%data['company_id']).dictresult()[0]['task_folder']
        task_folder="task_%s"%data['task_id']
        task_path='%s%s/%s/'%(cfg.task_path,folder,task_folder)

        if not os.path.exists(task_path):
            #en caso de no existir la carpeta, crea una nueva
            os.makedirs(task_path)
        else:
            #elimina evidencias de la carpeta antes de guardar las nuevas
            if data['status_id']==1:
                for the_file in os.listdir(task_path):
                    file_path = os.path.join(task_path, the_file)
                    if os.path.isfile(file_path):
                        os.unlink(file_path)


        if len(files_list)>0:
            for x in files_list:
                document_id=x.split("_")[1]
                file=frm[x]
                filename = secure_filename(file.filename)

                if os.path.exists("%s%s"%(task_path,filename)):
                    name,ext=os.path.splitext(filename)
                    filename="%s_%s%s"%(name,ev_cont,ext)
                    ev_cont+=1

                file.save(os.path.join(task_path, filename))
                file_size=os.path.getsize(os.path.join(task_path, filename))
                size_MB='%.4f'%(float(file_size)/1024/1024)

                db.query("""
                    update task.document
                    set file_path='%s',
                    file_name='%s',
                    loaded='now',
                    size='%s'
                    where task_id=%s
                    and document_id=%s
                """%(task_path,filename,size_MB,data['task_id'],document_id))
        db.query("""
            update task.task
            set
                resolved_date='now',
                status_id=2,
                comments='%s',
                last_updated='now',
                user_last_updated=%s,
                res_dec_by=%s
            where
                task_id=%s
            and company_id=%s
        """%(data['comments'],data['user_id'],data['user_id'],data['task_id'],data['company_id']))

        task_info=db.query("""
            select
                (select name from system.user where user_id=a.supervisor_id) as supervisor,
                a.name,
                (select name from system.user where user_id=a.assignee_id) as assignee,
                a.description,
                to_char(a.assignee_deadline,'DD-MM-YYYY HH24:MI:SS') as assignee_deadline,
                to_char(a.supervisor_deadline,'DD-MM-YYYY HH24:MI:SS') as supervisor_deadline,
                to_char(a.deadline,'DD-MM-YYYY HH24:MI:SS') as deadline,
                to_char(a.resolved_date,'DD-MM-YYYY HH24:MI:SS') as resolved_date,
                notify_admin
            from
                task.task a
            where a.task_id=%s
        """%data['task_id']).dictresult()[0]

        company_name=db.query("""
            select name from system.company where company_id=%s
        """%data['company_id']).dictresult()[0]
        task_info['company']=company_name['name']

        supervisor=db.query("""
            select email from system.user where user_id=(select supervisor_id from task.task
            where task_id=%s)
        """%data['task_id']).dictresult()[0]['email']
        message=db.query("""
            select * from template.generic_template where type_id=3
        """).dictresult()[0]
        task_info['link']=cfg.host
        task_info['mail_img']=cfg.mail_img
        msg=message['body'].format(**task_info)
        recipient=supervisor
        # GF.sendMail(message['subject'].format(**task_info),msg,supervisor)
        if task_info['notify_admin']==True: #si está indicado que se debe notificar al administrador al resolver la tarea
            admin=db.query("""
                select name,email from system.user
                where company_id=%s and user_type_id in (1,6)
            """%data['company_id']).dictresult()[0]
            task_info['admin']=admin['name']
            recipient=[supervisor,admin['email']]
            # message_admin=db.query("""
            #     select * from template.generic_template where type_id=3
            # """).dictresult()[0]
            # task_info['link']=cfg.host
            # task_info['mail_img']=cfg.mail_img
            # msg_admin=message_admin['body'].format(**task_info)
            # GF.sendMail(message_admin['subject'].format(**task_info),msg_admin,admin['email'])


        GF.sendMail(message['subject'].format(**task_info),msg,recipient)
        response['success']=True
        response['msg_response']='La tarea ha sido actualizada'
    except:
        response['success']=False
        response['msg_response']='Mal'
        exc_info=sys.exc_info()
        app.logger.info(traceback.format_exc(exc_info))
    return json.dumps(response)

@bp.route('/pauseResolveTask',methods=['GET','POST'])
@is_logged_in
def pauseResolveTask():
    response={}
    try:
        data=request.form.to_dict()
        files_list=eval(data['files_list'])
        files=request.files

        if len(files_list)>0: #if there are files to save
            #check if folder exists
            company_folder=db.query("""
                select task_folder from system.company where company_id=%s
            """%data['company_id']).dictresult()[0]['task_folder']
            path='%s%s/task_%s/'%(cfg.task_path,company_folder,data['task_id'])
            if not os.path.exists(path): #validate if folder exists
                os.makedirs(path) #if it doesn't exists, creates one
            for f in files_list:
                ev_cont=1
                document_id=f.split("_")[1] #get document_id

                #checks if it's already loaded
                is_loaded=db.query("""
                    select
                        (loaded<>'1900-01-01 00:00:00') as is_loaded,
                        file_path,
                        file_name
                    from task.document
                    where document_id=%s
                """%document_id).dictresult()[0]
                if is_loaded['is_loaded']==True: #if it's already loaded
                    remove_file=os.path.join(is_loaded['file_path'],is_loaded['file_name']) #get file path
                    if os.path.isfile(remove_file): #checks if it's a file
                        os.unlink(remove_file) #removes file
                    save_path=is_loaded['file_path'] #path where the file will be saved
                else: #if it hasn't been loaded
                    save_path=path #sets path where files are going to be saved

                file=files[f]
                filename=secure_filename(file.filename)
                if os.path.exists("%s%s"%(save_path,filename)): #if there's already a file with the same name
                    name,ext=os.path.splitext(filename)
                    filename="%s_%s%s"%(name,ev_cont,ext)
                    ev_cont+=1
                file.save(os.path.join(save_path,filename))
                file_size=os.path.getsize(os.path.join(save_path,filename))
                size_MB='%.4f'%(float(file_size)/1024/1024)
                db.query("""
                    update task.document
                    set file_path='%s',
                    file_name='%s',
                    loaded='now',
                    size='%s'
                    where task_id=%s
                    and document_id=%s
                """%(save_path,filename,size_MB,data['task_id'],document_id))
        comments=""
        if data['comments']!="": #checks if there are comments to be saved
            comments=" comments='%s',"%data['comments']
        db.query("""
            update task.task
            set
                status_id=6, %s
                last_updated='now',
                user_last_updated=%s
            where
                task_id=%s
            and company_id=%s
        """%(comments,data['user_id'],data['task_id'],data['company_id']))

        response['success']=True
        response['msg_response']='Los cambios realizados han sido guardados.'
    except:
        response['success']=False
        response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo.'
        exc_info=sys.exc_info()
        app.logger.info(traceback.format_exc(exc_info))
    return json.dumps(response)

@bp.route('/downloadEvidence/<document_id>', methods=['GET','POST'])
@is_logged_in
def downloadEvidence(document_id):
    response={}
    try:
        app.logger.info(document_id)
        doc_id=document_id.split("_")[1]
        evidence=db.query("""
            select
                file_name,
                file_path
            from
                task.document
            where
                document_id=%s
        """%doc_id.split(".")[0]).dictresult()

        path="%s%s"%(evidence[0]['file_path'],evidence[0]['file_name'])
        name=evidence[0]['file_name']

        return send_file(path,attachment_filename=name,as_attachment=True)
    except:
        response['success']=False
        response['msg_response']='Ocurrió un error'
        exc_info=sys.exc_info()
        app.logger.info(traceback.format_exc(exc_info))
        return response

@bp.route('/completeTask', methods=['GET','POST'])
@is_logged_in
def completeTask():
    response={}
    try:
        flag,data=GF.toDict(request.form,'post')
        if flag:
            db.query("""
                update task.task
                set
                    status_id=4,
                    last_updated='now',
                    user_last_updated=%s,
                    supervisor_comments='%s'
                where
                    task_id=%s
                and company_id=%s
            """%(data['user_id'],data['comments'],data['task_id'],data['company_id']))

            task_info=db.query("""
                select
                    (select name from system.user where user_id=a.supervisor_id) as supervisor,
                    a.name,
                    (select name from system.user where user_id=a.assignee_id) as assignee,
                    a.description,
                    to_char(a.assignee_deadline,'DD-MM-YYYY HH24:MI:SS') as assignee_deadline,
                    to_char(a.supervisor_deadline,'DD-MM-YYYY HH24:MI:SS') as supervisor_deadline,
                    to_char(a.deadline,'DD-MM-YYYY HH24:MI:SS') as deadline,
                    to_char(a.resolved_date,'DD-MM-YYYY HH24:MI:SS') as resolved_date,
                    supervisor_comments,
                    notify_admin
                from
                    task.task a
                where a.task_id=%s
            """%data['task_id']).dictresult()[0]
            company_name=db.query("""
                select name from system.company where company_id=%s
            """%data['company_id']).dictresult()[0]
            task_info['company']=company_name['name']
            assignee=db.query("""
                select email from system.user where user_id=(select assignee_id from task.task
                where task_id=%s)
            """%data['task_id']).dictresult()[0]['email']
            message=db.query("""
                select * from template.generic_template where type_id=32
            """).dictresult()[0]
            person_checking=db.query("""
                select name from system.user where user_id=%s
            """%data['user_id']).dictresult()[0]
            task_info['person_checking']=person_checking['name']
            task_info['link']=cfg.host
            task_info['mail_img']=cfg.mail_img
            msg=message['body'].format(**task_info)
            recipients=assignee
            if task_info['notify_admin']==True:
                admin=db.query("""
                    select email from system.user where company_id=%s and user_type_id in (1,6)
                """%data['company_id']).dictresult()[0]
                recipients=[assignee,admin['email']]


            GF.sendMail(message['subject'].format(**task_info),msg,recipients)

            response['success']=True
            response['msg_response']='La tarea ha sido cerrada.'
        else:
            response['success']=False
            response['msg_response']='Ocurrió un error al intentar procesar la información.'
    except:
        response['success']=False
        response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo.'
        exc_info=sys.exc_info()
        app.logger.info(traceback.format_exc(exc_info))
    return json.dumps(response)

@bp.route('/incompleteTask', methods=['GET','POST'])
@is_logged_in
def incompleteTask():
    response={}
    try:
        flag,data=GF.toDict(request.form, 'post')
        if flag:
            db.query("""
                update task.task
                set
                    status_id=1,
                    last_updated='now',
                    user_last_updated=%s,
                    supervisor_comments='%s'
                where
                    task_id=%s
                and company_id=%s
            """%(data['user_id'],data['comments'],data['task_id'],data['company_id']))
            db.query("""
                update task.document
                set loaded=default,
                file_name='',
                size=default
                where task_id=%s
            """%data['task_id'])
            task=db.query("""
                select
                    a.name as task,
                    b.name
                from
                    system.user b,
                    task.task a
                where
                    b.user_id=a.assignee_id
                and task_id=%s
            """%data['task_id']).dictresult()[0]
            response['success']=True
            response['msg_response']='La tarea %s ha sido asignada nuevamente a %s.'%(task['task'],task['name'])
            task_info=db.query("""
                select
                    (select name from system.user where user_id=a.supervisor_id) as supervisor,
                    a.name,
                    (select name from system.user where user_id=a.assignee_id) as assignee,
                    a.description,
                    to_char(a.assignee_deadline,'DD-MM-YYYY HH24:MI:SS') as assignee_deadline,
                    to_char(a.supervisor_deadline,'DD-MM-YYYY HH24:MI:SS') as supervisor_deadline,
                    to_char(a.deadline,'DD-MM-YYYY HH24:MI:SS') as deadline,
                    to_char(a.resolved_date,'DD-MM-YYYY HH24:MI:SS') as resolved_date,
                    supervisor_comments,
                    notify_admin
                from
                    task.task a
                where a.task_id=%s
            """%data['task_id']).dictresult()[0]
            company_name=db.query("""
                select name from system.company where company_id=%s
            """%data['company_id']).dictresult()[0]
            task_info['company']=company_name['name']

            assignee=db.query("""
                select email from system.user where user_id=(select assignee_id from task.task
                where task_id=%s)
            """%data['task_id']).dictresult()[0]['email']
            message=db.query("""
                select * from template.generic_template where type_id=33
            """).dictresult()[0]
            task_info['link']=cfg.host
            task_info['mail_img']=cfg.mail_img
            person_checking=db.query("""
                select name from system.user where user_id=%s
            """%data['user_id']).dictresult()[0]
            task_info['person_checking']=person_checking['name']
            msg=message['body'].format(**task_info)
            recipients=assignee
            if task_info['notify_admin']==True:
                admin=db.query("""
                    select email from system.user where company_id=%s and user_type_id in (1,6)
                """%data['company_id']).dictresult()[0]
                recipients=[assignee,admin['email']]

            GF.sendMail(message['subject'].format(**task_info),msg,recipients)

        else:
            response['success']=False
            response['msg_response']='Ocurrió un error al intentar procesar la información.'
    except:
        response['success']=False
        response['msg_response']='Ocurrió un error, favor de intentarlo nuevamente.'
        exc_info=sys.exc_info()
        app.logger.info(traceback.format_exc(exc_info))
    return json.dumps(response)

@bp.route('/declineTask', methods=['GET','POST'])
@is_logged_in
def declineTask():
    response={}
    try:
        flag,data=GF.toDict(request.form,'post')
        if flag:
            db.query("""
                update task.task
                set
                    status_id=3,
                    declining_cause='%s',
                    last_updated='now',
                    user_last_updated=%s,
                    res_dec_by=%s
                where
                    task_id=%s
                and company_id=%s
            """%(data['comments'],data['user_id'],data['user_id'],data['task_id'],data['company_id']))

            task_info=db.query("""
                select
                    (select name from system.user where user_id=a.supervisor_id) as supervisor,
                    a.name,
                    (select name from system.user where user_id=a.assignee_id) as assignee,
                    a.declining_cause,
                    a.description,
                    to_char(a.assignee_deadline,'DD-MM-YYYY HH24:MI:SS') as assignee_deadline,
                    to_char(a.supervisor_deadline,'DD-MM-YYYY HH24:MI:SS') as supervisor_deadline,
                    to_char(a.deadline, 'DD-MM-YYYY HH24:MI:SS') as deadline,
                    notify_admin
                from
                    task.task a
                where a.task_id=%s
            """%data['task_id']).dictresult()[0]
            task_info['link']=cfg.host
            task_info['mail_img']=cfg.mail_img
            company_name=db.query("""
                select name from system.company where company_id=%s
            """%data['company_id']).dictresult()[0]
            task_info['company']=company_name['name']

            recipient_list=[]
            if data['user_type_id']==3:
                recipient=db.query("""
                    select a.email
                    from
                        system.user a,
                        task.task b
                    where
                        a.user_id=b.supervisor_id
                    and b.task_id=%s
                """%data['task_id']).dictresult()[0]['email']
                # task_info=db.query("""
                #     select
                #         (select name from system.user where user_id=a.supervisor_id) as supervisor,
                #         a.name,
                #         (select name from system.user where user_id=a.assignee_id) as assignee,
                #         a.declining_cause,
                #         a.description,
                #         to_char(a.assignee_deadline,'DD-MM-YYYY HH24:MI:SS') as assignee_deadline,
                #         to_char(a.supervisor_deadline,'DD-MM-YYYY HH24:MI:SS') as supervisor_deadline,
                #         notify_admin
                #     from
                #         task.task a
                #     where a.task_id=%s
                # """%data['task_id']).dictresult()[0]

                message=db.query("""
                    select * from template.generic_template where type_id=6
                """).dictresult()[0]
                task_info['link']=cfg.host
                task_info['mail_img']=cfg.mail_img
                msg=message['body'].format(**task_info)
                recipient_list.append(recipient)
                # GF.sendMail(message['subject'].format(**task_info),msg,recipient)
                if task_info['notify_admin']==True:
                    admin=db.query("""
                        select email from system.user where company_id=%s and user_type_id in (1,6)
                    """%data['company_id']).dictresult()[0]
                    recipient_list.append(admin['email'])
            else:
                # task_info=db.query("""
                #     select
                #         (select name from system.user where user_id=a.supervisor_id) as supervisor,
                #         a.name,
                #         (select name from system.user where user_id=a.assignee_id) as assignee,
                #         a.declining_cause,
                #         a.description,
                #         to_char(a.assignee_deadline,'DD-MM-YYYY HH24:MI:SS') as assignee_deadline,
                #         to_char(a.supervisor_deadline,'DD-MM-YYYY HH24:MI:SS') as supervisor_deadline,
                #         to_char(a.deadline, 'DD-MM-YYYY HH24:MI:SS') as deadline
                #     from
                #         task.task a
                #     where a.task_id=%s
                # """%data['task_id']).dictresult()[0]
                task_info['link']=cfg.host
                task_info['mail_img']=cfg.mail_img

                assignee=db.query("""
                    select
                        email
                    from
                        system.user
                    where
                        user_id=(select assignee_id from task.task where task_id=%s)
                    and company_id=%s
                """%(data['task_id'],data['company_id'])).dictresult()[0]['email']
                message_assignee=db.query("""
                    select * from template.generic_template where type_id=16
                """).dictresult()[0]
                msg_assignee=message_assignee['body'].format(**task_info)
                recipient_list.append(assignee)
                # GF.sendMail(message_assignee['subject'].format(**task_info),msg_assignee,assignee)

                admin=db.query("""
                    select email,name from system.user where user_type_id in (1,6) and company_id=%s
                """%data['company_id']).dictresult()[0]
                task_info['admin']=admin['name']
                message_admin=db.query("""
                    select * from template.generic_template where type_id=17
                """).dictresult()[0]
                msg_admin=message_admin['body'].format(**task_info)
                recipient_list.append(admin['email'])
                # GF.sendMail(message_admin['subject'].format(**task_info),msg_admin,admin['email'])

            message_admin=db.query("""
                select * from template.generic_template where type_id=17
            """).dictresult()[0]
            msg_admin=message_admin['body'].format(**task_info)
            GF.sendMail(message_admin['subject'].format(**task_info),msg_admin,recipient_list)
            response['success']=True
            response['msg_response']='La tarea ha sido declinada.'
        else:
            response['success']=False
            response['msg_response']='Ocurrió un error al intentar procesar la información.'
    except:
        response['success']=False
        response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo.'
        exc_info=sys.exc_info()
        app.logger.info(traceback.format_exc(exc_info))
    return json.dumps(response)

@bp.route('/updateDeclinedTask', methods=['GET','POST'])
@is_logged_in
def updateDeclinedTask():
    response={}
    try:
        flag,data=GF.toDict(request.form,'post')
        if flag:
            if data['from']=='supervisor':
                if data['description']=="":
                    description=""
                else:
                    description="description='%s',"%data['description']

                db.query("""
                    update task.task
                    set assignee_id=%s,
                    supervisor_id=%s,
                    %s
                    status_id=1,
                    last_updated='now',
                    user_last_updated=%s
                    where task_id=%s
                """%(data['assignee_id'],data['supervisor_id'],description,data['user_id'],data['task_id']))
            else:
                db.query("""
                    update task.task
                    set assignee_id=%s,
                    supervisor_id=%s,
                    status_id=1,
                    deadline='%s',
                    assignee_deadline='%s',
                    supervisor_deadline='%s',
                    last_updated='now',
                    user_last_updated=%s
                    where task_id=%s
                """%(data['assignee_id'],data['supervisor_id'],data['deadline'],data['assignee_deadline'],data['supervisor_deadline'],data['user_id'],data['task_id']))

            task_info=db.query("""
                select
                    (select a.name from system.user a where a.user_id=assignee_id) as assignee,
                    (select a.email from system.user a where a.user_id=assignee_id) as assignee_mail,
                    (select a.name from system.user a where a.user_id=supervisor_id) as supervisor,
                    (select a.email from system.user a where a.user_id=supervisor_id) as supervisor_mail,
                    name,
                    description,
                    to_char(assignee_deadline, 'DD-MM-YYYY HH24:MI:SS') as assignee_deadline,
                    to_char(supervisor_deadline, 'DD-MM-YYYY HH24:MI:SS') as supervisor_deadline,
                    to_char(deadline, 'DD-MM-YYYY HH24:MI:SS') as deadline,
                    notify_admin,
                    company_id
                from
                    task.task
                where
                    task_id=%s
            """%data['task_id']).dictresult()[0]
            task_info['link']=cfg.host
            task_info['mail_img']=cfg.mail_img
            company_name=db.query("""
                select name from system.company where company_id=%s
            """%task_info['company_id']).dictresult()[0]
            task_info['company']=company_name['name']
            recipients=[]
            #Send mail to assignee
            msg_info_a=db.query("""
                select * from template.generic_template where type_id=34
            """).dictresult()[0]
            msg_a=msg_info_a['body'].format(**task_info)
            recipients.append(task_info['assignee_mail'])
            recipients.append(task_info['supervisor_mail'])
            if task_info['notify_admin']==True:
                admin=db.query("""
                    select email from system.user where company_id=%s and user_type_id in (1,6)
                """%task_info['company_id']).dictresult()[0]
                recipients.append(admin['email'])

            GF.sendMail(msg_info_a['subject'].format(**task_info),msg_a,recipients)

            #Send mail supervisor
            # msg_info_s=db.query("""
            #     select * from template.generic_template where type_id=19
            # """).dictresult()[0]
            # msg_s=msg_info_s['body'].format(**task_info)
            # GF.sendMail(msg_info_s['subject'].format(**task_info),msg_s,task_info['supervisor_mail'])
            response['success']=True
            response['msg_response']='La tarea ha sido reasignada.'
        else:
            response['success']=False
            response['msg_response']='Ocurrió un error al intentar obtener la información.'
    except:
        response['success']=False
        response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo.'
        exc_info=sys.exc_info()
        app.logger.info(traceback.format_exc(exc_info))
    return json.dumps(response)

@bp.route('/cancelTask', methods=['GET','POST'])
@is_logged_in
def cancelTask():
    response={}
    try:
        flag,data=GF.toDict(request.form,'post')
        if flag:
            db.query("""
                update task.task
                set status_id=5,
                last_updated='now',
                user_last_updated=%s
                where task_id=%s
            """%(data['user_id'],data['task_id']))

            task_info=db.query("""
                select
                    (select a.name from system.user a where a.user_id=assignee_id) as assignee,
                    (select a.name from system.user a where a.user_id=supervisor_id) as supervisor,
                    name,
                    description,
                    notify_admin,
                    company_id,
                    supervisor_id,
                    to_char(assignee_deadline,'DD-MM-YYYY HH24:MI:SS') as assignee_deadline,
                    to_char(supervisor_deadline,'DD-MM-YYYY HH24:MI:SS') as supervisor_deadline,
                    to_char(deadline, 'DD-MM-YYYY HH24:MI:SS') as deadline,
                    (select a.email from system.user a where a.user_id=assignee_id) as assignee_mail,
                    (select a.email from system.user a where a.user_id=supervisor_id) as supervisor_mail
                from
                    task.task
                where
                    task_id=%s
            """%data['task_id']).dictresult()[0]
            task_info['link']=cfg.host
            task_info['mail_img']=cfg.mail_img
            company_name=db.query("""
                select name from system.company where company_id=%s
            """%task_info['company_id']).dictresult()[0]
            task_info['company']=company_name['name']

            #send mail to assignee
            message_info_a=db.query("""
                select * from template.generic_template where type_id=20
            """).dictresult()[0]
            msg_a=message_info_a['body'].format(**task_info)
            recipients=[task_info['assignee_mail'],task_info['supervisor_mail']]
            if task_info['notify_admin']==True:
                admin=db.query("""
                    select user_id, email from system.user where company_id=%s and user_type_id in (1,6)
                """).dictresult()[0]
                if int(task_info['supervisor_id'])!=int(admin['user_id']):
                    recipients.append(admin['email'])

            GF.sendMail(message_info_a['subject'].format(**task_info),msg_a,recipients)

            # message_info_s=db.query("""
            #     select * from template.generic_template where type_id=20
            # """).dictresult()[0]
            # msg_s=message_info_s['body'].format(**task_info)
            # GF.sendMail(message_info_s['subject'].format(**task_info),msg_s,task_info['supervisor_mail'])

            response['success']=True
            response['msg_response']="La tarea ha sido cancelada."
        else:
            response['success']=False
            response['msg_response']="Ocurrió un error al intentar obtener la información."
    except:
        response['success']=False
        response['msg_response']="Ocurrió un error, favor de intentarlo de nuevo."
        exc_info=sys.exc_info()
        app.logger.info(traceback.format_exc(exc_info))
    return json.dumps(response)

@bp.route('/checkAssigneeTasks', methods=['GET','POST'])
@is_logged_in
def checkAssigneeTasks():
    response={}
    try:
        # app.logger.info(request.form)
        flag,data=GF.toDict(request.form,'post')
        if flag:
            response['overlaps']=False
            response['overlap_msg']=''
            tasks=db.query("""
                select name
                from task.task
                where assignee_deadline='%s 23:59:59'
                and company_id=%s
                and assignee_id=%s
                and status_id in (1,6)
            """%(data['assignee_deadline'],data['company_id'],data['assignee_id'])).dictresult()
            if tasks!=[]:
                response['overlaps']=True
                name=db.query("""
                    select name from system.user where user_id=%s
                """%data['assignee_id']).dictresult()
                msg=""
                for x in tasks:
                    msg+="<li>%s</li>"%GF.replaceStringHtml(x['name'])

                split_date=data['assignee_deadline'].split("-")
                new_name=GF.replaceStringHtml(name[0]['name'])
                new_date="%s-%s-%s"%(split_date[2],split_date[1],split_date[0])
                #response['overlap_msg']="El usuario %s tiene las siguientes tareas asignadas el d&iacute;a %s, ¿desea continuar?"%(new_name,new_date)
                response['overlap_msg']="El usuario %s tiene las siguientes tareas asignadas el d&iacute;a %s: <br><ul>%s</ul>&iquest;desea continuar?"%(new_name,new_date,msg)
            response['success']=True
        else:
            response['success']=False
            response['msg_response']='Ocurrió un error al intentar obtener los datos, favor de intentarlo de nuevo.'


    except:
        response['succes']=False
        response['msg_response']="Ocurrió un error, favor de intentarlo de nuevo."
        exc_info=sys.exc_info()
        app.logger.info(traceback.format_exc(exc_info))
    return json.dumps(response)


@bp.route('/doReport',methods=['GET','POST'])
@is_logged_in
def doReport():
    response={}
    try:
        flag,data=GF.toDict(request.form,'post')
        if flag:
            today=str(datetime.date.today())
            #assign dates in case from an to are empty
            if data['from']=="":
                if data['to']!="":
                    data['from']=data['to']
                else:
                    data['from']="%s-%s-01"%(today.split("-")[0],today.split("-")[1])
            if data['to']=="":
                if ['from']!="":
                    data['to']=data['from']
                else:
                    data['to']=today

            #validate that date from is before o equal to date to
            date_from=datetime.date(int(data['from'].split("-")[0]),int(data['from'].split("-")[1]),int(data['from'].split("-")[2]))
            date_to=datetime.date(int(data['to'].split("-")[0]),int(data['to'].split("-")[1]),int(data['to'].split("-")[2]))
            if date_from>date_to:
                #if date to is older than date from, assigns date from the same value as date to
                data['from']=data['to']

            list_list=[]


            if data['created_tasks']==True:
                created_tasks=db.query("""
                    select
                        a.name,
                        a.description,
                        to_char(a.deadline,'DD-MM-YYYY') as deadline,
                        (select name from system.user where user_id=a.supervisor_id) as supervisor,
                        to_char(a.supervisor_deadline,'DD-MM-YYYY') as supervisor_deadline,
                        (select name from system.user where user_id=a.assignee_id) as assignee,
                        to_char(a.assignee_deadline,'DD-MM-YYYY') as assignee_deadline,
                        to_char(a.created,'DD-MM-YYYY') as created,
                        (select name from system.user where user_id=a.created_by) as created_by
                    from
                        task.task a
                    where
                        company_id=%s
                    and created between '%s 00:00:00' and '%s 23:59:59'
                """%(data['company_id'],data['from'],data['to'])).dictresult()
                created_tasks_title=['Nombre','Descripción','Fecha límite','Supervisor','Fecha límite supervisor','Auxiliar','Fecha límite auxiliar','Fecha de creación','Creado por']
                created_tasks_d_order=['name','description','deadline','supervisor','supervisor_deadline','assignee','assignee_deadline','created','created_by']
                list_list.append({'title':created_tasks_title,'data':created_tasks,'sheet_name':'Tareas creadas','data_order':created_tasks_d_order,'color':"000040ff"})

            if data['assignee_tasks']==True:
                expired_assignee_tasks=db.query("""
                    select
                        a.name,
                        a.description,
                        to_char(a.deadline,'DD-MM-YYYY') as deadline,
                        (select name from system.user where user_id=a.supervisor_id) as supervisor,
                        to_char(a.supervisor_deadline,'DD-MM-YYYY') as supervisor_deadline,
                        (select name from system.user where user_id=a.assignee_id) as assignee,
                        to_char(a.assignee_deadline,'DD-MM-YYYY') as assignee_deadline,
                        to_char(a.created,'DD-MM-YYYY') as created,
                        (select name from system.user where user_id=a.created_by) as created_by,
                        to_char(a.last_updated, 'DD-MM-YYYY') as last_updated,
                        (select name from system.user where user_id=a.user_last_updated) as user_last_updated
                    from
                        task.task a
                    where
                        company_id=%s
                    and assignee_deadline between '%s 00:00:00' and '%s 23:59:59'
                    and status_id in (1,6)
                """%(data['company_id'],data['from'],data['to'])).dictresult()

                expired_assignee_tasks_title=['Nombre','Descripción','Fecha límite','Supervisor','Fecha límite supervisor','Auxiliar','Fecha límite auxiliar','Fecha de creación','Creado por','Actualizado por última vez','Actualizado por']
                expired_assignee_tasks_d_order=['name','description','deadline','supervisor','supervisor_deadline','assignee','assignee_deadline','created','created_by','last_updated','user_last_updated']
                list_list.append({'title':expired_assignee_tasks_title, 'data':expired_assignee_tasks,'sheet_name':'Tareas aux exp','data_order':expired_assignee_tasks_d_order,'color':"00ffd11a"})

            if data['supervisor_tasks']==True:
                expired_supervisor_tasks=db.query("""
                    select
                        a.name,
                        a.description,
                        to_char(a.deadline,'DD-MM-YYYY') as deadline,
                        (select name from system.user where user_id=a.supervisor_id) as supervisor,
                        to_char(a.supervisor_deadline,'DD-MM-YYYY') as supervisor_deadline,
                        (select name from system.user where user_id=a.assignee_id) as assignee,
                        to_char(a.assignee_deadline,'DD-MM-YYYY') as assignee_deadline,
                        to_char(a.created,'DD-MM-YYYY') as created,
                        (select name from system.user where user_id=a.created_by) as created_by,
                        to_char(a.last_updated, 'DD-MM-YYYY') as last_updated,
                        (select name from system.user where user_id=a.user_last_updated) as user_last_updated
                    from
                        task.task a
                    where
                        company_id=%s
                    and supervisor_deadline between '%s 00:00:00' and '%s 23:59:59'
                    and status_id in (1,6)
                """%(data['company_id'],data['from'],data['to'])).dictresult()

                expired_supervisor_tasks_title=['Nombre','Descripción','Fecha límite','Supervisor','Fecha límite supervisor','Auxiliar','Fecha límite auxiliar','Fecha de creación','Creado por','Actualizado por última vez','Actualizado por']
                expired_supervisor_tasks_d_order=['name','description','deadline','supervisor','supervisor_deadline','assignee','assignee_deadline','created','created_by','last_updated','user_last_updated']
                list_list.append({'title':expired_supervisor_tasks_title, 'data':expired_supervisor_tasks, 'sheet_name':'Tareas sup exp','data_order':expired_supervisor_tasks_d_order,'color':"00ff8000"})

            if data['admin_tasks']==True:
                expired_admin_tasks=db.query("""
                    select
                        a.name,
                        a.description,
                        to_char(a.deadline,'DD-MM-YYYY') as deadline,
                        (select name from system.user where user_id=a.supervisor_id) as supervisor,
                        to_char(a.supervisor_deadline,'DD-MM-YYYY') as supervisor_deadline,
                        (select name from system.user where user_id=a.assignee_id) as assignee,
                        to_char(a.assignee_deadline,'DD-MM-YYYY') as assignee_deadline,
                        to_char(a.created,'DD-MM-YYYY') as created,
                        (select name from system.user where user_id=a.created_by) as created_by,
                        to_char(a.last_updated, 'DD-MM-YYYY') as last_updated,
                        (select name from system.user where user_id=a.user_last_updated) as user_last_updated
                    from
                        task.task a
                    where
                        company_id=%s
                    and deadline between '%s 00:00:00' and '%s 23:59:59'
                    and status_id in (1,6)
                """%(data['company_id'],data['from'],data['to'])).dictresult()

                expired_admin_tasks_title=['Nombre','Descripción','Fecha límite','Supervisor','Fecha límite supervisor','Auxiliar','Fecha límite auxiliar','Fecha de creación','Creado por','Actualizado por última vez','Actualizado por']
                expired_admin_tasks_d_order=['name','description','deadline','supervisor','supervisor_deadline','assignee','assignee_deadline','created','created_by','last_updated','user_last_updated']
                list_list.append({'title':expired_admin_tasks_title, 'data':expired_admin_tasks,'sheet_name':'Tareas exp','data_order':expired_admin_tasks_d_order,'color':"00e62e00"})

            if data['resolved_tasks']==True:
                resolved_tasks=db.query("""
                    select
                        a.name,
                        a.description,
                        to_char(a.deadline,'DD-MM-YYYY') as deadline,
                        (select name from system.user where user_id=a.supervisor_id) as supervisor,
                        to_char(a.supervisor_deadline,'DD-MM-YYYY') as supervisor_deadline,
                        (select name from system.user where user_id=a.assignee_id) as assignee,
                        to_char(a.assignee_deadline,'DD-MM-YYYY') as assignee_deadline,
                        to_char(a.created,'DD-MM-YYYY') as created,
                        (select name from system.user where user_id=a.created_by) as created_by,
                        to_char(a.last_updated, 'DD-MM-YYYY') as last_updated,
                        (select name from system.user where user_id=a.user_last_updated) as user_last_updated,
                        to_char(a.resolved_date,'DD-MM-YYYY') as resolved_date,
                        a.comments
                    from
                        task.task a
                    where
                        company_id=%s
                    and resolved_date between '%s 00:00:00' and '%s 23:59:59'
                    and status_id = 2
                """%(data['company_id'],data['from'],data['to'])).dictresult()

                resolved_tasks_title=['Nombre','Descripción','Fecha límite','Supervisor','Fecha límite supervisor','Auxiliar','Fecha límite auxiliar','Fecha de creación','Creado por','Actualizado por última vez','Actualizado por','Fecha en que se resolvió','Comentarios']
                resolved_tasks_d_order=['name','description','deadline','supervisor','supervisor_deadline','assignee','assignee_deadline','created','created_by','last_updated','user_last_updated','resolved_date','comments']
                list_list.append({'title':resolved_tasks_title, 'data':resolved_tasks,'sheet_name':'Resueltas','data_order':resolved_tasks_d_order,'color':"0000e6b8"})

            if data['closed_tasks']==True:
                closed_tasks=db.query("""
                    select
                        a.name,
                        a.description,
                        to_char(a.deadline,'DD-MM-YYYY') as deadline,
                        (select name from system.user where user_id=a.supervisor_id) as supervisor,
                        to_char(a.supervisor_deadline,'DD-MM-YYYY') as supervisor_deadline,
                        (select name from system.user where user_id=a.assignee_id) as assignee,
                        to_char(a.assignee_deadline,'DD-MM-YYYY') as assignee_deadline,
                        to_char(a.created,'DD-MM-YYYY') as created,
                        (select name from system.user where user_id=a.created_by) as created_by,
                        to_char(a.last_updated, 'DD-MM-YYYY') as last_updated,
                        (select name from system.user where user_id=a.user_last_updated) as user_last_updated,
                        to_char(a.resolved_date,'DD-MM-YYYY') as resolved_date,
                        a.comments,
                        a.supervisor_comments
                    from
                        task.task a
                    where
                        company_id=%s
                    and last_updated between '%s 00:00:00' and '%s 23:59:59'
                    and status_id = 4
                """%(data['company_id'],data['from'],data['to'])).dictresult()

                closed_tasks_title=['Nombre','Descripción','Fecha límite','Supervisor','Fecha límite supervisor','Auxiliar','Fecha límite auxiliar','Fecha de creación','Creado por','Actualizado por última vez','Actualizado por','Fecha en que se resolvió','Comentarios','Comentarios del supervisor']
                closed_tasks_d_order=['name','description','deadline','supervisor','supervisor_deadline','assignee','assignee_deadline','created','created_by','last_updated','user_last_updated','resolved_date','comments','supervisor_comments']
                list_list.append({'title':closed_tasks_title, 'data':closed_tasks,'sheet_name':'Cerradas','data_order':closed_tasks_d_order,'color':"0000b300"})

            if data['declined_tasks']==True:
                declined_tasks=db.query("""
                    select
                        a.name,
                        a.description,
                        to_char(a.deadline,'DD-MM-YYYY') as deadline,
                        (select name from system.user where user_id=a.supervisor_id) as supervisor,
                        to_char(a.supervisor_deadline,'DD-MM-YYYY') as supervisor_deadline,
                        (select name from system.user where user_id=a.assignee_id) as assignee,
                        to_char(a.assignee_deadline,'DD-MM-YYYY') as assignee_deadline,
                        to_char(a.created,'DD-MM-YYYY') as created,
                        (select name from system.user where user_id=a.created_by) as created_by,
                        to_char(a.last_updated, 'DD-MM-YYYY') as last_updated,
                        (select name from system.user where user_id=a.user_last_updated) as user_last_updated,
                        a.declining_cause
                    from
                        task.task a
                    where
                        company_id=%s
                    and last_updated between '%s 00:00:00' and '%s 23:59:59'
                    and status_id = 3
                """%(data['company_id'],data['from'],data['to'])).dictresult()

                declined_tasks_title=['Nombre','Descripción','Fecha límite','Supervisor','Fecha límite supervisor','Auxiliar','Fecha límite auxiliar','Fecha de creación','Creado por','Actualizado por última vez','Actualizado por','Motivo para declinar']
                declined_tasks_d_order=['name','description','deadline','supervisor','supervisor_deadline','assignee','assignee_deadline','created','created_by','last_updated','user_last_updated','declining_cause']
                list_list.append({'title':declined_tasks_title, 'data':declined_tasks,'sheet_name':'Declinadas','data_order':declined_tasks_d_order,'color':"00cc00cc"})

            if data['canceled_tasks']==True:
                canceled_tasks=db.query("""
                    select
                        a.name,
                        a.description,
                        to_char(a.deadline,'DD-MM-YYYY') as deadline,
                        (select name from system.user where user_id=a.supervisor_id) as supervisor,
                        to_char(a.supervisor_deadline,'DD-MM-YYYY') as supervisor_deadline,
                        (select name from system.user where user_id=a.assignee_id) as assignee,
                        to_char(a.assignee_deadline,'DD-MM-YYYY') as assignee_deadline,
                        to_char(a.created,'DD-MM-YYYY') as created,
                        (select name from system.user where user_id=a.created_by) as created_by,
                        to_char(a.last_updated, 'DD-MM-YYYY') as last_updated,
                        (select name from system.user where user_id=a.user_last_updated) as user_last_updated,
                        a.declining_cause
                    from
                        task.task a
                    where
                        company_id=%s
                    and last_updated between '%s 00:00:00' and '%s 23:59:59'
                    and status_id = 5
                """%(data['company_id'],data['from'],data['to'])).dictresult()

                canceled_tasks_title=['Nombre','Descripción','Fecha límite','Supervisor','Fecha límite supervisor','Auxiliar','Fecha límite auxiliar','Fecha de creación','Creado por','Actualizado por última vez','Actualizado por','Motivo para declinar']
                canceled_tasks_d_order=['name','description','deadline','supervisor','supervisor_deadline','assignee','assignee_deadline','created','created_by','last_updated','user_last_updated','declining_cause']
                list_list.append({'title':canceled_tasks_title, 'data':canceled_tasks,'sheet_name':'Canceladas','data_order':canceled_tasks_d_order,'color':"00999999"})

            wb = Workbook()
            sheet_number = 0
            for ll in list_list:
                app.logger.info("------------Haciendo %s-----------------"%ll['sheet_name'])
                ws = wb.create_sheet("%s"%ll['sheet_name'],sheet_number)
                ws.sheet_properties.tabColor = ll['color']
                sheet_number+=1

                for i in range(1,len(ll['title'])+1):
                    ws.cell(row=1,column=i,value=ll['title'][i-1])
                    ws.cell(row=1,column=i).font=Font(name='Arial', size=12, bold=True)
                row=2
                if ll['data']!=[]:
                    for x in ll['data']:
                        for r in range(0,len(ll['data_order'])):
                            ws.cell(row=row,column=r+1,value=x[ll['data_order'][r]])
                            ws.cell(row=row,column=r+1).font=Font(name='Arial',size=11)
                        row+=1

            for w in wb.sheetnames:
                sheet=wb[w]
                for column_cells in sheet.columns:
                    length = max(len(GF.as_text(cell.value))+5 for cell in column_cells)
                    sheet.column_dimensions[column_cells[0].column].width = length

            fecha=str(datetime.datetime.today())
            fecha=fecha.replace(" ","_")
            fecha=fecha.replace(":","_")
            fecha=fecha.replace(".","_")



            if len(wb.sheetnames)>1:
                wb.remove(wb['Sheet'])
                wb.save('%sReporte_%s.xlsx'%(cfg.report_path,fecha))

                response['success']=True
                response['msg_response']="El reporte ha sido generado."
                response['filename']='/task/downloadReport/Reporte_%s.xlsx'%fecha

            else:
                response['msg_response']="No existe información en el periodo seleccionado."
                response['success']=False
        else:
            response['success']=False
            response['msg_response']="Ocurrió un error al intentar obtener los datos, favor de intentarlo de nuevo."

    except:
        response['success']=False
        response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo.'
        exc_info=sys.exc_info()
        app.logger.info(traceback.format_exc(exc_info))
    return json.dumps(response)

@bp.route('/downloadReport/<filename>', methods=['GET','POST'])
@is_logged_in
def downloadReport(filename):
    response={}
    try:
        path="%s%s"%(cfg.report_path,filename)
        name="%s"%filename
        return send_file(path,attachment_filename=name)

    except:
        response['success']=False
        response['msg_response']='Ocurrió un error'
        exc_info=sys.exc_info()
        app.logger.info(traceback.format_exc(exc_info))
        return response

@bp.route('/editTask', methods=['GET','POST'])
@is_logged_in
def editTask():
    response={}
    try:
        flag,data=GF.toDict(request.form,'post')
        if flag:

            deadline=time.strptime(data['deadline'],"%Y-%m-%d")
            supervisor_deadline=time.strptime(data['supervisor_deadline'],"%Y-%m-%d")
            assignee_deadline=time.strptime(data['assignee_deadline'],"%Y-%m-%d")
            if assignee_deadline<=supervisor_deadline and supervisor_deadline<=deadline:
                original_task=db.query("""
                    select
                        a.assignee_id, a.supervisor_id,
                        (select name from system.user where user_id=a.assignee_id) as assignee,
                        (select name from system.user where user_id=a.supervisor_id) as supervisor,
                        (select email from system.user where user_id=a.assignee_id) as assignee_mail,
                        (select email from system.user where user_id=a.supervisor_id) as supervisor_mail
                    from task.task a where a.task_id=%s
                """%data['task_id']).dictresult()[0]
                db.query("""
                    update task.task
                    set deadline='%s 23:59:59',
                    supervisor_deadline='%s 23:59:59',
                    assignee_deadline='%s 23:59:59',
                    assignee_id=%s,
                    supervisor_id=%s,
                    user_last_updated=%s,
                    last_updated='now'
                    where task_id=%s
                    and company_id=%s
                """%(data['deadline'],data['supervisor_deadline'],data['assignee_deadline'],data['assignee_id'],data['supervisor_id'],data['user_id'],data['task_id'],data['company_id']))

                task_info=db.query("""
                    select
                        to_char(a.deadline,'DD-MM-YYYY HH24:MI:SS') as deadline,
                        to_char(a.supervisor_deadline, 'DD-MM-YYYY HH24:MI:SS') as supervisor_deadline,
                        to_char(a.assignee_deadline, 'DD-MM-YYYY HH24:MI:SS') as assignee_deadline,
                        (select name from system.user where user_id=a.assignee_id) as assignee,
                        (select email from system.user where user_id=a.assignee_id) as assignee_mail,
                        (select name from system.user where user_id=a.supervisor_id) as supervisor,
                        (select email from system.user where user_id=a.supervisor_id) as supervisor_mail,
                        a.name,
                        a.description,
                        a.notify_admin
                    from
                        task.task a
                    where
                        a.task_id=%s
                """%data['task_id']).dictresult()[0]

                task_info['link']=cfg.host
                task_info['mail_img']=cfg.mail_img
                recipient_list=[]
                #send notification to new assignee
                company_name=db.query("""
                    select name from system.company where company_id=%s
                """%data['company_id']).dictresult()[0]
                task_info['company']=company_name['name']

                recipient_list.append(task_info['assignee_mail'])

                if int(data['assignee_id'])!=int(original_task['assignee_id']):

                    recipient_list.append(original_task['assignee_mail'])


                recipient_list.append(task_info['supervisor_mail'])

                if int(data['supervisor_id'])!=int(original_task['supervisor_id']):

                    recipient_list.append(original_task['supervisor_mail'])

                if task_info['notify_admin']==True:
                    admin_type=db.query("""
                        select user_id, user_type_id, name as admin, email as admin_mail  from system.user where company_id=%s and user_type_id in (1,6)
                    """%data['company_id']).dictresult()[0]
                    if int(admin_type['user_type_id'])==1:

                        recipient_list.append(admin_type['admin_mail'])
                    else:
                        if int(admin_type['user_id'])!=int(data['supervisor_id']):

                            recipient_list.append(admin_type['admin_mail'])


                message=db.query("""
                    select * from template.generic_template where type_id=25
                """).dictresult()[0]
                msg=message['body'].format(**task_info)
                GF.sendMail(message['subject'].format(**task_info),msg,recipient_list)

                response['success']=True
                response['msg_response']='La tarea ha sido actualizada.'
            else:
                response['success']=False
                response['msg_response']='La tarea no pudo ser actualizada, favor de revisar las fechas.'
        else:
            response['success']=False
            response['msg_response']='Ocurrió un error al intentar obtener los datos.'
    except:
        response['success']=False
        response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo.'
        exc_info=sys.exc_info()
        app.logger.info(traceback.format_exc(exc_info))
    return json.dumps(response)

@bp.route('/getAdminName', methods=['GET','POST'])
@is_logged_in
def getAdminName():
    response={}
    try:
        flag,data=GF.toDict(request.form,'post')
        if flag:
            if data['use']=='user':
                name=db.query("""
                    select name from system.user where user_id=%s
                """%data['user_id']).dictresult()
                if name!=[]:
                    response['name']=name[0]['name']
                    response['success']=True
                else:
                    response['success']=False
                    response['msg_response']="Ocurrió un error al intentar obtener los datos del usuario."
            elif data['use']=='company':
                name=db.query("""
                    select name from system.user where company_id=%s and user_type_id=1
                """%data['company_id']).dictresult()
                if name!=[]:
                    response['name']=name[0]['name']
                    response['success']=True
                else:
                    response['success']=False
                    response['msg_response']="Ocurrió un error al intentar obtener los datos del usuario."
        else:
            response['success']=False
            response['msg_response']="Ocurrió un error al intentar obtener los datos."
    except:
        response['success']=False
        response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo.'
        exc_info=sys.exc_info()
        app.logger.info(traceback.format_exc(exc_info))
    return json.dumps(response)

@bp.route('/getNotificationInfo', methods=['GET','POST'])
@is_logged_in
def getNotificationInfo():
    response={}
    try:
        flag,data=GF.toDict(request.form,'post')
        if flag:
            assignee=db.query("""
                select
                    a.user_id,
                    a.name
                from
                    system.user a,
                    task.task b
                where
                    b.task_id=%s
                and a.user_id=b.assignee_id
            """%data['task_id']).dictresult()[0]
            select_list=[]
            if data['user_type_id']==2 or data['user_type_id']==3: #if sender is supervisor
                supervisor=db.query("""
                    select
                        a.user_id,
                        a.name
                    from
                        system.user a,
                        task.task b
                    where
                        a.user_id=b.supervisor_id
                    and b.task_id=%s
                """%data['task_id']).dictresult()[0]
                admin=db.query("""
                    select
                        name,
                        user_id
                    from
                        system.user
                    where
                        company_id=%s
                    and user_type_id in (1,6)
                """%data['company_id']).dictresult()[0]
                if data['user_type_id']==2:
                    select_list.append({
                        'name':'%s - Aux'%assignee['name'],
                        'user_id':'%s'%assignee['user_id']
                    })
                    response['msg_from']=supervisor['name']
                elif data['user_type_id']==3:
                    select_list.append({
                        'name':'%s - Sup'%supervisor['name'],
                        'user_id':'%s'%supervisor['user_id']
                    })
                    response['msg_from']=assignee['name']
                if int(admin['user_id'])!=int(supervisor['user_id']):
                    select_list.append({
                        'name':'%s - Admin'%admin['name'],
                        'user_id':'%s'%admin['user_id']
                    })
                    if data['user_type_id']==2:
                        select_list.append({
                            'name':'%s - Aux, CC: %s - Admin'%(assignee['name'],admin['name']),
                            'user_id':'%s,%s'%(assignee['user_id'],admin['user_id'])
                        })
                    else:
                        select_list.append({
                            'name':'%s - Sup, CC: %s - Admin'%(supervisor['name'],admin['name']),
                            'user_id':'%s,%s'%(supervisor['user_id'],admin['user_id'])
                        })



            elif data['user_type_id']==6: #if sender is supervisor-admin
                supervisor=db.query("""
                    select
                        a.user_id,
                        a.name
                    from
                        system.user a,
                        task.task b
                    where
                        a.user_id=b.supervisor_id
                    and b.task_id=%s
                """%data['task_id']).dictresult()[0]
                select_list.append({
                    'name':'%s - Aux'%assignee['name'],
                    'user_id':'%s'%assignee['user_id']
                })
                if int(supervisor['user_id'])!=int(data['user_id']):
                    select_list.append({
                        'name':'%s - Sup'%supervisor['name'],
                        'user_id':'%s'%supervisor['user_id']
                    })
                    select_list.append({
                        'name':'%s - Aux, %s - Sup'%(assignee['name'],supervisor['name']),
                        'user_id':'%s,%s'%(assignee['user_id'],supervisor['user_id'])
                    })
                admin=db.query("""
                    select name from system.user where company_id=%s and user_type_id=6
                """%data['company_id']).dictresult()[0]
                response['msg_from']=admin['name']

            elif data['user_type_id']==1: #if sender is admin
                supervisor=db.query("""
                    select
                        a.user_id,
                        a.name
                    from
                        system.user a,
                        task.task b
                    where
                        a.user_id=b.supervisor_id
                    and b.task_id=%s
                """%data['task_id']).dictresult()[0]
                admin=db.query("""
                    select
                        name,
                        user_id
                    from
                        system.user
                    where
                        company_id=%s
                    and user_type_id=1
                """%data['company_id']).dictresult()[0]
                select_list.append({
                    'name':'%s - Aux'%assignee['name'],
                    'user_id':'%s'%assignee['user_id']
                })
                select_list.append({
                    'name':'%s - Sup'%supervisor['name'],
                    'user_id':'%s'%supervisor['user_id']
                })
                select_list.append({
                    'name':'%s - Aux, CC: %s - Sup'%(assignee['name'],supervisor['name']),
                    'user_id':'%s,%s'%(assignee['user_id'],supervisor['user_id'])
                })
                response['msg_from']=admin['name']

            elif data['user_type_id']==4:
                supervisor=db.query("""
                    select
                        a.user_id,
                        a.name
                    from
                        system.user a,
                        task.task b
                    where
                        a.user_id=b.supervisor_id
                    and b.task_id=%s
                """%data['task_id']).dictresult()[0]
                admin=db.query("""
                    select
                        name,
                        user_id
                    from
                        system.user
                    where
                        company_id=%s
                    and user_type_id in (1,6)
                """%data['company_id']).dictresult()[0]
                select_list.append({
                    'name':'%s - Aux'%assignee['name'],
                    'user_id':'%s'%assignee['user_id']
                })
                select_list.append({
                    'name':'%s - Sup'%supervisor['name'],
                    'user_id':'%s'%supervisor['user_id']
                })
                select_list.append({
                    'name':'%s - Admin'%admin['name'],
                    'user_id':'%s'%admin['user_id']
                })
                select_list.append({
                    'name':'%s - Aux, CC: %s - Sup'%(assignee['name'],supervisor['name']),
                    'user_id':'%s,%s'%(assignee['user_id'],supervisor['user_id'])
                })
                select_list.append({
                    'name':'%s - Aux, CC: %s - Admin'%(assignee['name'],admin['name']),
                    'user_id':'%s,%s'%(assignee['user_id'],admin['user_id'])
                })
                if int(supervisor['user_id'])!=int(admin['user_id']):
                    select_list.append({
                        'name':'%s - Sup, CC: %s - Admin'%(supervisor['name'],admin['name']),
                        'user_id':'%s,%s'%(supervisor['user_id'],admin['user_id'])
                    })
                consultant=db.query("""
                    select name from system.user where user_id=%s
                """%data['user_id']).dictresult()[0]

                response['msg_from']=consultant['name']

            response['select_list']=select_list
            response['success']=True

        else:
            response['success']=False
            response['msg_response']="Ocurrió un error al intentar obtener los datos."
    except:
        response['success']=False
        response['msg_response']="Ocurrió un error, favor de intentarlo de nuevo."
        exc_info=sys.exc_info()
        app.logger.info(traceback.format_exc(exc_info))
    return json.dumps(response)


@bp.route('/sendNotification', methods=['GET','POST'])
@is_logged_in
def sendNotification():
    response={}
    try:
        flag,data=GF.toDict(request.form,'post')
        if flag:
            template=db.query("""
                select * from template.generic_template where type_id=28
            """).dictresult()[0]
            task_info=db.query("""
                select
                    a.task_id,
                    a.name,
                    a.description,
                    company_id,
                    (select name from system.user where user_id=a.assignee_id) as assignee,
                    (select name from system.user where user_id=a.supervisor_id) as supervisor,
                    to_char(a.deadline, 'DD-MM-YYYY HH24:MI:SS') as deadline,
                    to_char(a.supervisor_deadline, 'DD-MM-YYYY HH24:MI:SS') as supervisor_deadline,
                    to_char(a.assignee_deadline, 'DD-MM-YYYY HH24:MI:SS') as assignee_deadline
                from
                    task.task a
                where
                    a.task_id=%s
            """%data['task_id']).dictresult()[0]
            task_info['msg']=data['message']
            msg_from=db.query("""
                select name,email from system.user where user_id=%s
            """%data['msg_from']).dictresult()[0]
            recipients=data['msg_to'].split(",")
            if len(recipients)==1:
                msg_to=db.query("""
                    select name as msg_to, email from system.user where user_id=%s
                """%data['msg_to']).dictresult()[0]
                mail_recipients=[msg_to['email']]
            else:
                msg_to=db.query("""
                    select name as msg_to from system.user where user_id=%s
                """%recipients[0]).dictresult()[0]
                cc=db.query("""
                    select email from system.user where user_id in (%s)
                """%data['msg_to']).dictresult()
                mail_recipients=[]
                for c in cc:
                    mail_recipients.append(c['email'])
            mail_recipients.append(msg_from['email'])

            task_info['msg_to']=msg_to['msg_to']
            task_info['msg_from']=msg_from['name']
            task_info['msg']=data['message'].encode('utf-8')
            task_info['link']=cfg.host
            task_info['mail_img']=cfg.mail_img
            company_name=db.query("""
                select name from system.company where company_id=%s
            """%task_info['company_id']).dictresult()[0]
            task_info['company']=company_name['name']
            message=template['body'].format(**task_info)

            GF.sendMail(template['subject'].format(**task_info),message,mail_recipients)
            task_notification={
                'task_id':data['task_id'],
                'msg_from':data['msg_from'],
                'msg_to':data['msg_to'],
                'message':data['message'].encode('utf-8'),
                'send_date':'now'
            }
            db.insert("task.notification",task_notification)
            response['success']=True
            response['msg_response']="La notificación ha sido enviada."

        else:
            response['success']=False
            response['msg_response']="Ocurrió un error al intentar obtener los datos."
    except:
        response['success']=False
        response['msg_response']="Ocurrió un error, favor de intentarlo de nuevo."
        exc_info=sys.exc_info()
        app.logger.info(traceback.format_exc(exc_info))
    return json.dumps(response)

@bp.route('/getNotificationHistory', methods=['GET','POST'])
@is_logged_in
def getNotificationHistory():
    response={}
    try:
        flag,data=GF.toDict(request.form,'post')
        if flag:
            notif=db.query("""
                select *,
                    to_char(send_date,'DD-MM-YYYY HH24:MI:SS') as send_date
                from task.notification where task_id=%s order by notification_id desc
            """%data['task_id']).dictresult()
            if notif!=[]:
                response['has_messages']=True
                divs=''
                for x in notif:
                    msg_from=db.query("""
                        select name from system.user where user_id = %s
                    """%x['msg_from']).dictresult()[0]
                    msg_to_list=db.query("""
                        select name from system.user where user_id in (%s)
                    """%x['msg_to']).dictresult()
                    msg_to=', '.join(e['name'] for e in msg_to_list)
                    divs+='''
                        <div class="card card-mail"><div class="card-header card-header-mail"><h6><b>De:</b> %s</h6><h6><b>A:</b> %s</h6></div><div class="card-body card-body-mail"><span class="span-mail-date">Enviado:%s</span><hr class="hr-mail-date" /><p class="card-text">%s</p></div></div>
                    '''%(msg_from['name'],msg_to,x['send_date'],x['message'])
                response['mail_divs']=divs
            else:
                response['has_messages']=False
            response['success']=True
        else:
            response['success']=False
            response['msg_response']="Ocurrió un error al intentar obtener los datos, favor de intentarlo de nuevo."
    except:
        response['success']=False
        response['msg_response']="Ocurrió un error, favor de intentarlo de nuevo."
        exc_info=sys.exc_info()
        app.logger.info(traceback.format_exc(exc_info))
    return json.dumps(response)

@bp.route('/getUploadTasksButtons',methods=['GET','POST'])
@is_logged_in
def getUploadTasksButtons():
    response={}
    try:
        flag,data=GF.toDict(request.form,'post')
        if flag:
            buttons='''
                <a href="/task/downloadAssigneeList/{company_id}.xlsx" role="button" class="btn btn-outline-primary" data-toggle="tooltip" title="Descargar lista de auxiliares" style="margin-left:3%; margin-right:2.5%; width:44.5%;" target="_blank">Lista de auxiliares</a><a href="/task/downloadSupervisorList/{company_id}.xlsx" role="button" class="btn btn-outline-primary" data-toggle="tooltip" title="Descargar lista de supervisores" style="margin-left:2.5%; margin-right:3%; width:44.5%;" target="_blank">Lista supervisores</a>
            '''.format(**data)
            response['success']=True
            response['buttons']=buttons
        else:
            response['success']=False
            response['msg_response']='Ocurrió un error al intentar obtener los datos.'
    except:
        response['success']=False
        response['msg_response']="Ocurrió un error, favor de intentarlo más tarde."
        exc_info=sys.exc_info()
        app.logger.info(traceback.format_exc(exc_info))
    return json.dumps(response)

@bp.route('/downloadAssigneeList/<company_id>',methods=['GET','POST'])
@is_logged_in
def downloadAssigneeList(company_id):
    response={}
    try:
        app.logger.info(company_id)
        assignees=db.query("""
            select name, login from system.user where company_id=%s and user_type_id=3
        """%company_id.split(".")[0]).dictresult()
        wb = Workbook()
        sheet_number = 0
        ws = wb.create_sheet("Auxiliares")
        ws.cell(row=1,column=1,value='Nombre').font=Font(name='Arial', size=12, bold=True)
        ws.cell(row=1,column=2,value='Usuario').font=Font(name='Arial', size=12, bold=True)
        row=2
        for x in assignees:
            ws.cell(row=row, column=1, value=x['name']).font=Font(name='Arial',size=11)
            ws.cell(row=row, column=2, value=x['login']).font=Font(name='Arial',size=11)
            row+=1

        for column_cells in ws.columns:
            length = max(len(GF.as_text(cell.value))+5 for cell in column_cells)
            ws.column_dimensions[column_cells[0].column].width = length

        wb.remove(wb['Sheet'])
        wb.save('%sAuxiliares.xlsx'%cfg.report_path)

        path="%sAuxiliares.xlsx"%(cfg.report_path)
        name='Auxiliares.xlsx'

        return send_file(path,attachment_filename=name,as_attachment=True)
    except:
        response['success']=False
        response['msg_response']="Ocurrió un error, favor de intentarlo de nuevo."
        exc_info=sys.exc_info()
        app.logger.info(traceback.format_exc(exc_info))
        return json.dumps(response)

@bp.route('/downloadSupervisorList/<company_id>',methods=['GET','POST'])
@is_logged_in
def downloadSupervisorList(company_id):
    response={}
    try:
        app.logger.info(company_id)
        supervisors=db.query("""
            select name, login from system.user where company_id=%s and user_type_id=2
        """%company_id.split(".")[0]).dictresult()
        wb = Workbook()
        sheet_number = 0
        ws = wb.create_sheet("Supervisores")
        ws.cell(row=1,column=1,value='Nombre').font=Font(name='Arial', size=12, bold=True)
        ws.cell(row=1,column=2,value='Usuario').font=Font(name='Arial', size=12, bold=True)
        row=2
        for x in supervisors:
            ws.cell(row=row, column=1, value=x['name']).font=Font(name='Arial',size=11)
            ws.cell(row=row, column=2, value=x['login']).font=Font(name='Arial',size=11)
            row+=1

        for column_cells in ws.columns:
            length = max(len(GF.as_text(cell.value))+5 for cell in column_cells)
            ws.column_dimensions[column_cells[0].column].width = length

        wb.remove(wb['Sheet'])
        wb.save('%sSupervisores.xlsx'%cfg.report_path)

        path="%sSupervisores.xlsx"%(cfg.report_path)
        name='Supervisores.xlsx'

        return send_file(path,attachment_filename=name,as_attachment=True)
    except:
        response['success']=False
        response['msg_response']="Ocurrió un error, favor de intentarlo de nuevo."
        exc_info=sys.exc_info()
        app.logger.info(traceback.format_exc(exc_info))
        return json.dumps(response)

@bp.route('/downloadTaskFormat', methods=['GET','POST'])
@is_logged_in
def downloadTaskFormat():
    response={}
    try:
        headers=['No. tarea','Nombre','Descripción','Fecha de vencimiento','Supervisor','Fecha de supervisor','Auxiliar','Fecha auxiliar','Notificar a administrador','Tarea recurrente','Frecuencia','Nombre evidencia','Descripción evidencia','Tipo de documento']
        wb = Workbook()
        ws = wb.create_sheet("Tareas")
        column=1
        for x in headers:
            ws.cell(row=1,column=column,value=x).font=Font(name='Arial', size=11)
            column+=1



        columna=['Columna','No. tarea', 'Nombre','Descripción','Fecha de vencimiento','Supervisor','Fecha supervisor','Auxiliar','Fecha auxiliar','Notificar administrador','Tarea recurrente','Frecuencia (meses)','Nombre evidencia*','Descripción evidencia*','Tipo de documento*']
        descripcion=['Descripción','Número de la tarea, es un consecutivo que iniciará desde 1.','Nombre de la tarea.','Descripción de la tarea (opcional).','Fecha de vencimiento de la tarea.','Supervisor que estará a cargo de revisar la tarea.','Fecha de vencimiento de la tarea para el supervisor.','Auxiliar a quien será asignada la tarea.','Fecha de vencimiento de la tarea para el auxiliar.','Indica si se debe enviar una notificación al administrador en cuanto la tarea sea resuelta.','Indica si desea que la tarea se genere de manera automática cada determinado tiempo.','Frecuencia en meses con la que se creará la tarea seleccionada como recurrente.','Nombre de evidencia solicitada.','Descripción de evidencia solicitada.','Tipo de documento de la evidencia.']
        valores=['Valores aceptados','1,2,3, etc.','Texto','Texto','Fecha formato dd-mm-yyyy','Usuario con el que inicia sesión','Fecha formato dd-mm-yyyy','Usuario con el que inicia sesión','Fecha formato dd-mm-yyyy','si/no','si/no','1-12| En caso de indicar "no" en recurrente, se deberá llenar esta celda con un 0 (cero).','Texto','Texto','xml, pdf, excel, texto, zip, power_point']
        evidence='*Cuando existe más de una evidencia por tarea, se deberán agregar en las filas siguientes, dejando vacíos las columnas correspondientes a la tarea, y agregando solo el número de la tarea (sin importar el número de evidencias que se agreguen, el número de la tarea siempre es el mismo por cada tarea), el nombre de la evidencia, la descripción de la evidencia y el tipo de documento.'

        ws2 = wb.create_sheet("Instrucciones")
        instrucciones=[columna,descripcion,valores]
        column_number=1
        for i in instrucciones:
            row=1
            for x in i:
                if row==1:
                    ws2.cell(row=row,column=column_number,value=x).font=Font(name='Arial', size=12, color='FFFFFF')
                else:
                    ws2.cell(row=row,column=column_number,value=x).font=Font(name='Arial', size=11)
                row+=1
            column_number+=1

        ws2['A1'].fill=PatternFill(start_color="8f93c7", end_color="8f93c7", fill_type = "solid")
        ws2['B1'].fill=PatternFill(start_color="8f93c7", end_color="8f93c7", fill_type = "solid")
        ws2['C1'].fill=PatternFill(start_color="8f93c7", end_color="8f93c7", fill_type = "solid")


        sheetnames=wb.sheetnames
        for sheet in sheetnames:
            for column_cells in wb[sheet].columns:
                length = max(len(GF.as_text(cell.value))+5 for cell in column_cells)
                wb[sheet].column_dimensions[column_cells[0].column].width = length

        ws2.merge_cells(start_row=19, start_column=1, end_row=22, end_column=2)
        ws2['A19']=evidence
        wrap_alignment = Alignment(wrapText=True)
        ws2['A19'].alignment = wrap_alignment
        ws2['A19'].font=Font(name='Arial', size=10)

        wb.remove(wb['Sheet'])
        wb.save('%sFormato_tareas.xlsx'%cfg.report_path)
        path="%sFormato_tareas.xlsx"%(cfg.report_path)
        name='Formato_tareas.xlsx'

        return send_file(path,attachment_filename=name,as_attachment=True)

    except:
        response['success']=False
        response['msg_response']="Ocurrió un error, favor de intentarlo de nuevo."
        exc_info=sys.exc_info()
        app.logger.info(traceback.format_exc(exc_info))
        return json.dumps(response)

@bp.route('/uploadTasks', methods=['GET','POST'])
@is_logged_in
def uploadTasks():
    response={}
    try:
        data=request.form.to_dict()
        files=request.files
        file_path=cfg.report_path
        file=files[data['file_name']]
        filename = secure_filename(file.filename)
        file.save(os.path.join(file_path, filename))

        read_file=load_workbook(os.path.join(file_path,filename))

        # sheet_task=read_file[0]

        active_sheet=read_file['Tareas']
        sheet_rows=active_sheet.rows
        # app.logger.info(act_sheet.rows)

        cont=0
        keys=[]
        dicts=[]
        error=""

        row_error=[]
        key_dicts=[]

        inserted_tasks=[]

        for x in sheet_rows:
            if cont==0:
                for y in x:
                    enc_key=y.value.encode('utf-8')
                    key=GF.replaceString(enc_key)
                    keys.append(key.lower())
            else:
                row_d={}
                cont2=0
                for y in x:
                    row_d[keys[cont2]]=y.value
                    cont2+=1
                dicts.append(row_d)
            cont+=1



        for d in dicts:
            nd={}
            try:
                #intenta armar el diccionario, si lo logra armar, quiere decir que corresponde a una fila con datos de la tarea
                #va a caer en la excepción cuando intente hacer el split o un encode
                #busca el id del usuario correspondiente al supervisor

                sup_id=db.query("""
                    select user_id from system.user where login='%s' and company_id=%s
                """%(d['supervisor'].lower(),data['company_id'])).dictresult()

                if sup_id!=[]:
                    # nd['supervisor_login']=d['supervisor']
                    nd['supervisor_id']=sup_id[0]['user_id']
                    #busca el id del usuario correspondiente al auxiliar
                    aux_id=db.query("""
                        select user_id from system.user where login='%s' and company_id=%s
                    """%(d['auxiliar'].lower(),data['company_id'])).dictresult()

                    if aux_id!=[]:
                        # nd['assignee_login']=d['auxiliar']

                        nd['assignee_id']=aux_id[0]['user_id']
                        nd['deadline']="%s-%s-%s"%(d['fecha_vencimiento'].split("-")[2],d['fecha_vencimiento'].split("-")[1],d['fecha_vencimiento'].split("-")[0])
                        nd['supervisor_deadline']="%s-%s-%s"%(d['fecha_supervisor'].split("-")[2],d['fecha_supervisor'].split("-")[1],d['fecha_supervisor'].split("-")[0])
                        nd['assignee_deadline']="%s-%s-%s"%(d['fecha_auxiliar'].split("-")[2],d['fecha_auxiliar'].split("-")[1],d['fecha_auxiliar'].split("-")[0])
                        #comparar fechas - se debe validar que: deadline>=supervisor_deadline>=assignee_deadline
                        deadline=datetime.datetime(int(nd['deadline'].split("-")[0]),int(nd['deadline'].split("-")[1]),int(nd['deadline'].split("-")[2]))
                        sup_deadline=datetime.datetime(int(nd['supervisor_deadline'].split("-")[0]),int(nd['supervisor_deadline'].split("-")[1]),int(nd['supervisor_deadline'].split("-")[2]))
                        aux_deadline=datetime.datetime(int(nd['assignee_deadline'].split("-")[0]),int(nd['assignee_deadline'].split("-")[1]),int(nd['assignee_deadline'].split("-")[2]))
                        if deadline>=sup_deadline>=aux_deadline:

                            nd['recurrent']=d['tarea_recurrente']
                            nd['task_number']=int(d['no._tarea'])
                            nd['notify_admin']=d['notificar_administrador']
                            nd['name']=d['nombre'].encode('utf-8')

                            if d['descripcion']!=None:
                                nd['description']=d['descripcion'].encode('utf-8')
                            else:
                                nd['description']=d['nombre'].encode('utf-8')

                            if d['tarea_recurrente'].lower()=='si':
                                nd['recurrent']=True
                                nd['recurrent_frequency']=int(d['frecuencia'])
                            else:
                                nd['recurrent']=False
                                nd['recurrent_frequency']=0

                            nd['documents']=[]
                            if d['nombre_evidencia']!=None:
                                evidence={}
                                doc_type_id=db.query("""
                                    select document_type_id from task.document_type
                                    where document_type='%s'
                                """%d['tipo_de_documento'].replace("_"," ").upper()).dictresult()
                                if doc_type_id!=[]:
                                    evidence['document_type_id']=doc_type_id[0]['document_type_id']
                                    evidence['name']=d['nombre_evidencia'].encode('utf-8')
                                    # evidence['document_type']=d['tipo_de_documento']
                                    if d['descripcion_evidencia']!=None:
                                        evidence['description']=d['descripcion_evidencia'].encode('utf-8')
                                    if len(evidence)>0:
                                        nd['documents'].append(evidence)
                                else:
                                    error="Error en la evidencia '{nombre_evidencia}', no se encontró el tipo de documento '{tipo_de_documento}'".format(**d)
                                    row_error.append(error)
                            key_dicts.append(nd)
                        else:
                            tarea=d['nombre']
                            error="Favor de revisar las fechas de la tarea {nombre}, tomando en cuenta que la fecha límite debe ser mayor o igual a la fecha de supervisor, y la fecha de supervisor debe ser mayor o igual a la fecha del auxiliar.".format(**d)
                            row_error.append(error)

                    else:
                        error="Error en la tarea '{nombre}': El auxiliar '{auxiliar}' no se encuentra registrado.".format(**d)
                        row_error.append(error)
                else:
                    error="Error en la tarea '{nombre}': El supervisor '{supervisor}' no se encuentra registrado.".format(**d)
                    row_error.append(error)


            except:
                #cuando cae en la excepción, es porque no logra armar el diccionario de la tarea, osea solo contiene datos de evidencias
                if d['nombre_evidencia']!=None:
                    for kd in key_dicts:
                        if int(d['no._tarea'])==int(kd['task_number']):
                            doc_type_id=db.query("""
                                select document_type_id from task.document_type
                                where document_type='%s'
                            """%d['tipo_de_documento'].replace("_"," ").upper()).dictresult()
                            if doc_type_id!=[]:
                                evidence={
                                    'name':d['nombre_evidencia'].encode('utf-8'),
                                    'document_type_id':doc_type_id[0]['document_type_id'],
                                    'description':''
                                }
                                if d['descripcion_evidencia']!=None:
                                    evidence['description']=d['descripcion_evidencia'].encode('utf-8')
                                kd['documents'].append(evidence)
                            else:
                                error="Error en la evidencia '{nombre_evidencia}', no se encontró el tipo de documento '{tipo_de_documento}'".format(**d)
                                row_error.append(error)

                            break

        #guardar tareas
        if key_dicts!=[]:
            for dic in key_dicts:
                dic['company_id']=data['company_id']

                dic['status_id']=1
                dic['user_last_updated']=data['user_id']
                dic['created_by']=data['user_id']
                new_task=db.insert("task.task",dic)
                inserted_tasks.append(new_task['task_id']) #agrega a una lista las tareas insertadas
                if dic['documents']!=[]:
                    for d in dic['documents']:
                        d['task_id']=new_task['task_id']
                        a=db.insert("task.document",d)

            response['success']=True
            response['has_error_msg']=False

            if row_error!=[]:
                response['has_error_msg']=True
                error_html="Ocurri&oacute; un error al intentar guardar los siguientes registros:<br><ul>"
                for re in row_error:
                    error_html+="<li>%s</li>"%re
                error_html+="</ul>"
                response['error_msg']=error_html
        else:
            response['success']=False
            response['msg_response']="No se guardó ninguna tarea."
        app.logger.info("before subprocess")
        __location__ = os.path.realpath(os.path.join(os.getcwd(), os.path.dirname(__file__)))
        sub=subprocess.Popen(['python', os.path.join(__location__,'mail_subprocess.py'),str(inserted_tasks),str(data['company_id'])],bufsize=0, executable=None, stdin=None, stdout=None, stderr=None, preexec_fn=None, close_fds=False, shell=False, cwd=None, env=None, universal_newlines=False, startupinfo=None, creationflags=0)
        app.logger.info(sub)
        app.logger.info("after subprocess")

    except:
        response['success']=False
        response['msg_response']="Ocurrió un error, favor de intentarlo de nuevo."
        exc_info=sys.exc_info()
        app.logger.info(traceback.format_exc(exc_info))
    return json.dumps(response)

@bp.route('/hideTask', methods=['GET','POST'])
@is_logged_in
def hideTask():
    response={}
    try:
        flag,data=GF.toDict(request.form,'post')
        if flag:
            db.query("""
                update task.task
                set hidden=True
                where task_id=%s
            """%data['task_id'])
            response['success']=True
            response['msg_response']='La tarea ha quedado oculta.'
        else:
            response['success']=False
            response['msg_response']='Ocurrió un error al intentar obtener los datos de la tarea.'
    except:
        response['success']=False
        response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo más tarde.'
        exc_info=sys.exc_info()
        app.logger.info(traceback.format_exc(exc_info))
    return json.dumps(response)

@bp.route('/showTask', methods=['GET','POST'])
@is_logged_in
def showTask():
    response={}
    try:
        flag,data=GF.toDict(request.form,'post')
        if flag:
            db.query("""
                update task.task
                set hidden=False
                where task_id=%s
            """%data['task_id'])
            response['success']=True
            response['msg_response']='La tarea ha quedado visible.'
        else:
            response['success']=False
            response['msg_response']='Ocurrió un error al intentar obtener los datos de la tarea.'
    except:
        response['success']=False
        response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo más tarde.'
        exc_info=sys.exc_info()
        app.logger.info(traceback.format_exc(exc_info))
    return json.dumps(response)

@bp.route('/getTodayTasks', methods=['GET','POST'])
@is_logged_in
def getTodayTasks():
    response={}
    try:
        flag,data=GF.toDict(request.form,'post')
        deadline=""
        #auxiliar
        if int(data['user_type_id'])==3:
            deadline=" assignee_deadline "
        #supervisor
        elif int(data['user_type_id'])==2:
            deadline=" supervisor_deadline "
        else:
            deadline=" deadline "
        today=str(datetime.datetime.today()).split(" ")[0]
        tasks=db.query("""
            select task_id, name from task.task where %s between '%s 00:00:00' and '%s 23:59:59'
            and company_id=%s and (supervisor_id=%s or assignee_id=%s) and status_id in (1,6) order by name asc
        """%(deadline,today,today,data['company_id'],data['user_id'],data['user_id'])).dictresult()

        task_html=""
        if tasks!=[]:
            for t in tasks:
                name=hex(int(t['task_id'])*cfg.factor_tt)
                html='<a class="today-task" href="#" name="%s" data-toggle="tooltip" title="Click para ir a la tarea">%s</a><br>'%(name,t['name'])
                task_html+=html
            response['tasks']=task_html
        else:
            response['tasks']="<h2>No tienes tareas pendientes para hoy.</h2>"
        response['success']=True


    except:
        response['success']=False
        response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo más tarde.'
        exc_info=sys.exc_info()
        app.logger.info(traceback.format_exc(exc_info))
    return json.dumps(response)
