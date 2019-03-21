#!/usr/bin/env python
#--*-- coding: utf-8 --*--
from flask import Flask, render_template, flash, redirect, url_for, session, request, logging, Blueprint, g, send_file
from passlib.hash import sha256_crypt
from functools import wraps
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.datastructures import ImmutableMultiDict
from werkzeug.utils import secure_filename
from .db_connection import getDB
db = getDB()
from .auth import is_logged_in
import json
import traceback
from flask_mail import Mail, Message
from . import generic_functions
GF = generic_functions.GenericFunctions()
from flask import current_app as app
import sys
import time
import os
import random
import app_config as cfg
import datetime
import time
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Color, colors, PatternFill, Border, Alignment
from openpyxl.cell import Cell

bp = Blueprint('project', __name__, url_prefix='/project')

@bp.route('/saveProject', methods=['GET','POST'])
@is_logged_in
def saveProject():
    response = {}
    try:
        if request.method == 'POST':
            flag,data = GF.toDict(request.form, 'post')
            if flag:
                if data['new_project_id']==-1:
                    data['created_by'] = data['user_id']
                    data['status_id'] = 1
                    data['resolved_date'] = '1900-01-01 00:00:00'
                    data['last_updated_by'] = data['user_id']
                    new_project = db.insert("task.project",data)
                    response['new']=True

                    response['project_id']=new_project['project_id']
                else:
                    db.query("""
                        update task.project
                        set name='%s',
                        description='%s',
                        deadline='%s 00:00:00',
                        last_updated_by=%s,
                        last_updated=now()
                        where project_id=%s
                    """%(data['name'],data['description'],data['deadline'],data['user_id'],data['new_project_id']))
                    response['new']=False
                response['success']=True
            else:
                response['success'] = False
                response['msg_response'] = 'Ocurrió un error al intentar procesar la información. Favor de intentarlo de nuevo.'

        else:
            response['success'] = False
            response['msg_response'] = 'Ocurrió un error, favor de intentarlo de nuevo.'

    except:
        response['success'] = False
        response['msg_response'] = 'Ocurrió un error.'
        exc_info = sys.exc_info()
        app.logger.info(traceback.format_exc(exc_info))
    return json.dumps(response)

@bp.route('/getProjectInfo', methods=['GET','POST'])
@is_logged_in
def getProjectInfo():
    response = {}
    try:
        if request.method == 'POST':
            flag,data = GF.toDict(request.form, 'post')
            if flag:
                project=db.query("""
                    select
                        a.name,
                        a.description,
                        to_char(a.deadline,'DD-MM-YYYY HH:MI:SS') as deadline,
                        (select name from system.user where user_id=a.created_by) as created_by,
                        to_char(a.last_updated, 'DD-MM-YYYY HH:MI:SS') as last_updated,
                        (select name from system.user where user_id=a.last_updated_by) as last_updated_by
                    from
                        task.project a
                    where
                        a.project_id=%s
                """%data['project_id']).dictresult()
                if project!=[]:
                    html="""
                        <p><b>Nombre: </b>{name}<br><b>Descripción: </b>{description}<br><b>Fecha de vencimiento: </b>{deadline}<br><b>Creado por: </b>{created_by}<br><b>Actualizado por última vez: </b>{last_updated}<br><b>Actualizado por: </b>{last_updated_by}</p>
                    """.format(**project[0])
                    response['html'] = html
                    response['success'] = True
                    response['project_id'] = data['project_id']
                else:
                    response['success'] = False
                    response['msg_response'] = 'El proyecto no fue encontrado. Favor de intentarlo de nuevo.'
            else:
                response['success'] = False
                response['msg_response'] = 'Ocurrió un error al intentar procesar la información. Favor de intentarlo de nuevo.'

        else:
            response['success'] = False
            response['msg_response'] = 'Ocurrió un error, favor de intentarlo de nuevo.'

    except:
        response['success'] = False
        response['msg_response'] = 'Ocurrió un error.'
        exc_info = sys.exc_info()
        app.logger.info(traceback.format_exc(exc_info))
    return json.dumps(response)

@bp.route('/getProjectList', methods=['GET','POST'])
@is_logged_in
def getProjectList():
    response={}
    try:
        if request.method=='POST':
            flag,data=GF.toDict(request.form,'post')
            if flag:
                projects=db.query("""
                    select project_id, name
                    from task.project
                    where company_id=%s and status_id in (1,6)
                """%data['company_id']).dictresult()
                projects.append({'project_id':-1,'name':'Ninguno'})
                response['success']=True
                response['data']=projects
            else:
                response['success']=False
                response['msg_response']='Ocurrió un error al intentar obtener los datos. Favor de intentarlo de nuevo.'
        else:
            response['success']=False
            response['msg_response']='Ocurrió un error. Favor de intentarlo de nuevo.'
    except:
        response['success']=False
        response['msg_response']='Ocurrió un error'
        exc_info=sys.exc_info()
        app.logger.info(traceback.format_exc(exc_info))
    return json.dumps(response)

@bp.route('/getProjects', methods=['GET','POST'])
@is_logged_in
def getProjects():
    response={}
    try:
        if request.method=='POST':
            user_type_id=int(request.form['user_type_id'])
            user_id=int(request.form['user_id'])
            company_id=int(request.form['company_id'])
            first=request.form['first']
            filter=json.loads(request.form['filter'])

            if first=="true":
                response['first']=True
                oldest_project=db.query("""
                    select to_char(deadline,'YYYY-MM-DD') as deadline,
                    to_char(deadline,'DD-MM-YYYY') as deadline_return
                    from task.project
                    where company_id=%s
                    and status_id=1
                    order by deadline asc
                    limit 1
                """%company_id).dictresult()

                newest_project=db.query("""
                    select
                        to_char(deadline,'YYYY-MM-DD') as deadline,
                        to_char(deadline,'DD-MM-YYYY') as deadline_return
                    from task.project
                    where company_id=%s
                    and status_id=1
                    order by deadline desc limit 1
                """%company_id).dictresult()
                status=" and a.status_id=1"
                created_by=""
                search=""
                if oldest_project!=[]:
                    date_from=oldest_project[0]['deadline']
                    date_to=newest_project[0]['deadline']
                    response['date_from']=oldest_project[0]['deadline']
                    response['date_to']=newest_project[0]['deadline']
                else:
                    today=datetime.datetime.now()
                    date_from='%s-%s-01'%(today.year,str(today.month).zfill(2))
                    date_to='%s-%s-%s'%(today.year,str(today.month).zfill(2),str(today.day).zfill(2))
                    response['date_from']=False
                    response['date_to']=False
                date_filter=" and a.deadline between '%s 00:00:00' and '%s 23:59:59'"%(date_from,date_to)
            else:
                response['first']=False
                date_from=filter['from']
                date_to=filter['to']
                status=""
                created_by=""
                search=""
                if filter['status_id']!=-1:
                    status=" and a.status_id=%s"%filter['status_id']
                if filter['created_by']!=-1:
                    created_by=" and a.created_by=%s"%filter['created_by']
                if int(filter['date_type'])==1:
                    date_filter=" and a.created between '%s 00:00:00' and '%s 23:59:59'"%(date_from,date_to)
                else:
                    date_filter=" and a.deadline between '%s 00:00:00' and '%s 23:59:59'"%(date_from,date_to)
                if filter['search']!="":
                    search=" and a.name ilike '%%%s%%'"%filter['search']

            projects=db.query("""
                select
                    a.project_id,
                    a.name,
                    a.description,
                    (select name from system.user where user_id=a.created_by) as created_by,
                    to_char(a.created,'DD-MM-YYYY') as created,
                    to_char(a.deadline,'DD-MM-YYYY') as deadline,
                    (select count(task_id) from task.task where project_id=a.project_id) as tasks,
                    b.description as status,
                    a.status_id
                from
                    task.project a,
                    task.status b
                where
                    company_id=%s
                and a.status_id=b.status_id
                %s %s %s %s
                order by a.name
                offset %s limit %s
            """%(company_id,status,created_by,date_filter,search,int(request.form['start']),int(request.form['length']))).dictresult()


            total=db.query("""
                select
                    count(*)
                from
                    task.project a
                where
                    a.company_id=%s
                    %s %s %s %s
            """%(company_id,status,created_by,date_filter,search)).dictresult()

            response['data']=projects
            response['recordsTotal']=total[0]['count']
            response['recordsFiltered']=total[0]['count']
            response['success']=True

        else:
            response['success']=False
            response['msg_response']='Ocurrió un error, cargue la página nuevamente.'
    except:
        response['success']=False
        response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo.'
        exc_info=sys.exc_info()
        app.logger.info(traceback.format_exc(exc_info))
    return json.dumps(response)

@bp.route('/getCreatedBy', methods=['GET','POST'])
@is_logged_in
def getCreatedBy():
    response={}
    try:
        flag,data=GF.toDict(request.form,'post')
        if flag:
            users=db.query("""
                select
                    user_id,
                    name
                from
                    system.user
                where
                    company_id=%s
                and enabled in (1,3)
                order by name
            """%data['company_id']).dictresult()
            users.append({'user_id':-1,'name':'Todos'})
            response['data']=users
            response['success']=True
        else:
            response['success']=False
            response['msg_response']='Ocurrió un error, favor de cargar la página nuevamente.'
    except:
        response['success']=False
        response['msg_response']='Ocurrió un error, favor de intentarlo nuevamente.'
        exc_info=sys.exc_info()
        app.logger.info(traceback.format_exc(exc_info))
    return json.dumps(response)

@bp.route('/getProjectTasks', methods=['GET','POST'])
@is_logged_in
def getProjectTasks():
    response={}
    try:
        if request.method=='POST':
            company_id=int(request.form['company_id'])
            project_id=int(request.form['project_id'])
            tasks=db.query("""
                select
                    a.task_id,
                    a.name,
                    (select name from system.user where user_id=a.supervisor_id) as supervisor,
                    (select name from system.user where user_id=a.assignee_id) as assignee,
                    b.description as status,
                    to_char(a.deadline,'DD-MM-YYYY') as deadline,
                    a.status_id
                from
                    task.status b,
                    task.task a
                where
                    a.project_id=%s
                and a.status_id=b.status_id
                and a.company_id=%s
                order by a.name
                offset %s limit %s
            """%(project_id,company_id,int(request.form['start']),int(request.form['length']))).dictresult()

            total_tasks=db.query("""
                select count(a.task_id)
                from task.task a
                where a.project_id=%s
                and a.company_id=%s
            """%(project_id,company_id)).dictresult()

            response['data']=tasks
            response['recordsTotal']=total_tasks[0]['count']
            response['recordsFiltered']=total_tasks[0]['count']
            response['success']=True

        else:
            response['success']=False
            response['msg_response']='Ocurrió un error, favor de intentarlo nuevamente.'
    except:
        response['success']=False
        response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo.'
        exc_info=sys.exc_info()
        app.logger.info(traceback.format_exc(exc_info))
    return json.dumps(response)

@bp.route('/getSearchedProjectTasks', methods=['GET','POST'])
@is_logged_in
def getSearchedProjectTasks():
    response={}
    try:
        if request.method=='POST':
            filters=json.loads(request.form['filters'])
            company_id=int(request.form['company_id'])
            user_id=int(request.form['user_id'])
            user_type_id=int(request.form['user_type_id'])
            if int(filters['date_type'])==2:
                date_filter=" and a.deadline between '%s 00:00:00' and '%s 23:59:59' "%(filters['from'],filters['to'])
            else:
                date_filter=" and a.created between '%s 00:00:00' and '%s 23:59:59' "%(filters['from'],filters['to'])
            search_filter=""
            if filters['search']!="":
                search_filter=" and a.name ilike '%%%s%%'"%filters['search']
            status_filter=""
            if int(filters['status_id'])!=-1:
                status_filter=" and a.status_id=%s"%filters['status_id']
            tasks=db.query("""
                select
                    a.task_id,
                    a.name,
                    to_char(a.deadline,'DD-MM-YYYY') as deadline,
                    (select name from system.user where user_id=a.supervisor_id) as supervisor,
                    (select name from system.user where user_id=a.assignee_id) as assignee,
                    b.description as status
                from
                    task.status b,
                    task.task a
                where
                    a.company_id=%s
                and a.status_id=b.status_id
                %s %s %s
                order by a.name
                offset %s limit %s
            """%(company_id,status_filter,date_filter,search_filter,int(request.form['start']),int(request.form['length']))).dictresult()
            task_total=db.query("""
                select count(a.task_id)
                from task.task a
                where a.company_id=%s
                and a.status_id=%s %s %s
            """%(company_id,filters['status_id'],date_filter,search_filter)).dictresult()
            response['data']=tasks
            response['recordsTotal']=task_total[0]['count']
            response['recordsFiltered']=task_total[0]['count']
            response['success']=True
        else:
            response['success']=False
            response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo.'
    except:
        response['success']=False
        response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo.'
        exc_info=sys.exc_info()
        app.logger.info(traceback.format_exc(exc_info))
    return json.dumps(response)

#revisa si la tarea ya se encuentra agregada a algún proyecto
@bp.route('/checkTaskProject', methods=['GET','POST'])
@is_logged_in
def checkTaskProject():
    response={}
    try:
        flag,data=GF.toDict(request.form,'post')
        if flag:
            #obtiene id del proyecto de la tarea seleccionada
            task_project=db.query("""
                select project_id from task.task
                where task_id=%s
            """%data['task_id']).dictresult()
            # app.logger.info(type(task_project[0]['project_id']))
            if task_project!=[]:
                if int(task_project[0]['project_id'])!=-1:
                    project_name=db.query("""
                        select name from task.project where project_id=%s
                    """%task_project[0]['project_id']).dictresult()
                    response['success']=True
                    response['needs_confirm']=True
                    response['msg_response']='La tarea se encuentra asignada al proyecto %s, ¿Desea continuar?'%project_name[0]['name']
                else:
                    response['success']=True
                    response['needs_confirm']=False
            else:
                response['success']=False
                response['msg_response']='Ocurrió un problema al obtener los datos de la tarea seleccionada, favor de intentarlo de nuevo más tarde.'
        else:
            response['success']=False
            response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo.'
    except:
        response['success']=False
        response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo más de tarde.'
        exc_info=sys.exc_info()
        app.logger.info(traceback.format_exc(exc_info))
    return json.dumps(response)

@bp.route('/addTaskToProject', methods=['GET','POST'])
@is_logged_in
def addTaskToProject():
    response={}
    try:
        flag,data=GF.toDict(request.form,'post')
        if flag:
            #buscar proyecto al que está asignada la tarea.
            task_project=db.query("""
                select project_id
                from task.task
                where task_id=%s
            """%data['task_id']).dictresult()
            if task_project!=[]:
                if int(task_project[0]['project_id'])==int(data['project_id']):
                    response['success']=False
                    response['msg_response']='La tarea ya se encuentra agregada al proyecto.'
                else:
                    db.query("""
                        update task.task
                        set project_id=%s,
                        last_updated=now(),
                        user_last_updated=%s
                        where task_id=%s
                    """%(data['project_id'],data['user_id'],data['task_id']))
                    db.query("""
                        update task.project
                        set last_updated=now(),
                        last_updated_by=%s
                        where project_id=%s
                    """%(data['user_id'],data['project_id']))
                    response['success']=True
                    response['msg_response']='La tarea ha sido agregada al proyecto.'
            else:
                response['success']=False
                response['msg_response']='Ocurrió un problema al intentar obtener los datos del proyecto, favor de intentarlo de nuevo más tarde.'
        else:
            response['success']=False
            response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo más tarde.'
    except:
        response['success']=False
        response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo más tarde.'
        exc_info=sys.exc_info()
        app.logger.info(traceback.format_exc(exc_info))
    return json.dumps(response)

@bp.route('/removeTaskFromProject', methods=['GET','POST'])
@is_logged_in
def removeTaskFromProject():
    response={}
    try:
        flag,data=GF.toDict(request.form, 'post')
        if flag:
            db.query("""
                update task.task
                set project_id=-1
                where task_id=%s
            """%data['task_id'])
            db.query("""
                update task.project
                set last_updated=now(),
                last_updated_by=%s
                where project_id=%s
            """%(data['user_id'],data['project_id']))
            response['success']=True
            response['msg_response']='La tarea ha sido eliminada del proyecto.'
        else:
            response['success']=False
            response['msg_response']='Ocurrió un error al intentar obtener los datos de la tarea, favor de intentarlo de nuevo más tarde.'
    except:
        response['success']=False
        response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo más tarde.'
        exc_info=sys.exc_info()
        app.logger.info(traceback.format_exc(exc_info))
    return json.dumps(response)

@bp.route('/deleteProject', methods=['GET','POST'])
@is_logged_in
def deleteProject():
    response={}
    try:
        flag,data=GF.toDict(request.form,'post')
        if flag:
            db.query("""
                update task.task
                set project_id=-1
                where project_id=%s
            """%data['project_id'])
            db.query("""
                delete from task.project where project_id=%s
            """%data['project_id'])
            response['success']=True
            response['msg_response']='El proyecto ha sido eliminado.'
        else:
            response['success']=False
            response['msg_response']='Ocurrió un error al intentar obtener los datos del proyecto, favor de intentarlo de nuevo.'
    except:
        response['success']=False
        response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo más tarde.'
        exc_info=sys.exc_info()
        app.logger.info(traceback.format_exc(exc_info))
    return json.dumps(response)

@bp.route('/getRecipientsInfo', methods=['GET','POST'])
@is_logged_in
def getRecipientsInfo():
    response={}
    try:
        flag,data=GF.toDict(request.form,'post')
        if flag:
            project_user=db.query("""
                select created_by from task.project where project_id=%s
            """%data['project_id']).dictresult()
            task_users=db.query("""
                select
                    supervisor_id,
                    assignee_id,
                    created_by
                from
                    task.task
                where
                    project_id=%s
            """%data['project_id']).dictresult()
            users=[]
            users.append(str(int(project_user[0]['created_by'])))
            for x in task_users:
                users.append(str(int(x['supervisor_id'])))
                users.append(str(int(x['assignee_id'])))
                users.append(str(int(x['created_by'])))
            unique_users=list(set(users))
            user_list=','.join(unique_users)

            users_info=db.query("""
                select user_id, name from system.user where user_id in (%s)
            """%user_list).dictresult()
            html=''
            div='''
                <div class="custom-control custom-checkbox" style="margin-left:10px;">
                    <input type="checkbox" class="custom-control-input" id="check_user_{user_id}" name="user_id_{user_id}">
                    <label class="custom-control-label" for="check_user_{user_id}">{name}</label>
                </div>
            '''

            for u in users_info:
                html+=div.format(**u)

            sender=db.query("""
                select user_id, name from system.user where user_id=%s
            """%data['user_id']).dictresult()
            response['sender']=sender[0]['name']
            response['success']=True
            response['divs']=html
        else:
            response['success']=False
            response['msg_response']='Ocurrió un error al intentar obtener los datos de los usuarios, favor de intentarlo más tarde.'

    except:
        response['success']=False
        response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo más tarde.'
        exc_info=sys.exc_info()
        app.logger.info(traceback.format_exc(exc_info))
    return json.dumps(response)


@bp.route('/sendProjectNotification', methods=['GET','POST'])
@is_logged_in
def sendProjectNotification():
    response={}
    try:
        flag,data=GF.toDict(request.form,'post')
        if flag:
            template=db.query("""
                select * from template.generic_template where type_id=35
            """).dictresult()
            project_info=db.query("""
                select
                    a.name,
                    a.description,
                    (select name from system.user where user_id=a.created_by) as created_by,
                    to_char(a.deadline,'DD-MM-YYYY HH:MI:SS') as deadline,
                    (select name from system.company where company_id=a.company_id) as company
                from
                    task.project a
                where
                    a.project_id=%s
            """%data['project_id']).dictresult()[0]
            task_names=db.query("""
                select
                    name
                from
                    task.task
                where project_id=%s
            """%data['project_id']).dictresult()
            html_task_names=""
            # app.logger.info(task_names)
            if task_names!=[]:
                html_task_names=' '.join('<li>%s</li>'%tn['name'] for tn in task_names)

            sender=db.query("""
                select name, email from system.user where user_id=%s
            """%data['sender']).dictresult()
            recipients=[]
            for d in data:
                if d[:8]=='user_id_':
                    if data[d]==True:
                        recipients.append(d.split("user_id_")[1])

            recipients_ids=','.join(recipients)
            recipient_mails=db.query("""
                select email from system.user where user_id in (%s)
            """%recipients_ids).dictresult()
            recipients_mail_list=[]
            for rm in recipient_mails:
                recipients_mail_list.append(rm['email'])
            project_info['msg_from']=sender[0]['name']
            project_info['msg']=data['message'].encode('utf-8')
            project_info['link']=cfg.host
            project_info['mail_img']=cfg.mail_img
            project_info['tasks']=html_task_names
            message=template[0]['body'].format(**project_info)
            # app.logger.info(message)
            GF.sendMail(template[0]['subject'].format(**project_info),message,recipients_mail_list)
            project_notification={
                'project_id':data['project_id'],
                'msg_from':data['sender'],
                'msg_to':recipients_ids,
                'message':data['message'].encode('utf-8'),
                'send_date':'now'
            }
            db.insert("task.project_notification",project_notification)
            response['success']=True
            response['msg_response']='La notificación ha sido enviada.'
        else:
            response['success']=False
            response['msg_response']='Ocurrió un error al intentar obtener los datos del proyecto, favor de intentarlo de nuevo más tarde.'
    except:
        response['success']=False
        response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo más tarde.'
        exc_info=sys.exc_info()
        app.logger.info(traceback.format_exc(exc_info))
    return json.dumps(response)


@bp.route('/getProjectNotificationHistory', methods=['GET','POST'])
@is_logged_in
def getProjectNotificationHistory():
    response={}
    try:
        flag,data=GF.toDict(request.form,'post')
        if flag:
            notif=db.query("""
                select *,
                to_char(send_date,'DD-MM-YYYY HH:MI:SS') as send_date
                from task.project_notification
                where project_id=%s
                order by project_notification_id desc
            """%data['project_id']).dictresult()
            if notif!=[]:
                response['has_messages']=True
                divs=''
                for x in notif:
                    msg_from=db.query("""
                        select name from system.user where user_id=%s
                    """%x['msg_from']).dictresult()[0]
                    msg_to_list=db.query("""
                        select name from system.user where user_id in (%s)
                    """%x['msg_to']).dictresult()
                    msg_to=', '.join(e['name'] for e in msg_to_list)
                    divs+='''
                        <div class="card card-mail"><div class="card-header card-header-mail"><h6><b>De:</b> %s</h6><h6><b>A:</b> %s</h6></div><div class="card-body card-body-mail"><span class="span-mail-date">Enviado: %s</span><hr class="hr-mail-date" /><p class="card-text">%s</p></div></div>
                    '''%(msg_from['name'],msg_to,x['send_date'],x['message'])
                response['messages']=divs
            else:
                response['has_messages']=False
            response['success']=True
        else:
            response['success']=False
            response['msg_response']='Ocurrió un error al intentar obtener los datos del proyecto, favor de intentarlo de nuevo.'
    except:
        response['success']=False
        response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo más tarde.'
        exc_info=sys.exc_info()
        app.logger.info(traceback.format_exc(exc_info))
    return json.dumps(response)

@bp.route('/getProjectSummary', methods=['GET','POST'])
@is_logged_in
def getProjectSummary():
    response={}
    try:
        flag, data=GF.toDict(request.form,'post')
        if flag:
            project_info=db.query("""
                select
                    a.name,
                    a.description,
                    to_char(a.created,'DD-MM-YYYY HH24:MI:SS') as created,
                    (select name from system.user where user_id=a.created_by) as created_by,
                    to_char(a.deadline,'DD-MM-YYYY HH24:MI:SS') as deadline
                from
                    task.project a
                where
                    a.project_id=%s
            """%data['project_id']).dictresult()[0]

            closed_tasks=db.query("""
                select
                    a.name,
                    to_char(a.assignee_deadline,'DD-MM-YYYY HH24:MI:SS') as assignee_deadline,
                    to_char(a.supervisor_deadline,'DD-MM-YYYY HH24:MI:SS') as supervisor_deadline,
                    to_char(a.deadline,'DD-MM-YYYY HH24:MI:SS') as deadline,
                    (select name from system.user where user_id=a.assignee_id) as assignee,
                    (select name from system.user where user_id=a.supervisor_id) as supervisor,
                    to_char(a.resolved_date,'DD-MM-YYYY HH24:MI:SS') as resolved_date
                from
                    task.task a
                where
                    project_id=%s
                and status_id=4
            """%data['project_id']).dictresult()

            resolved_tasks=db.query("""
                select
                    a.name,
                    to_char(a.assignee_deadline,'DD-MM-YYYY HH24:MI:SS') as assignee_deadline,
                    to_char(a.supervisor_deadline,'DD-MM-YYYY HH24:MI:SS') as supervisor_deadline,
                    to_char(a.deadline,'DD-MM-YYYY HH24:MI:SS') as deadline,
                    (select name from system.user where user_id=a.assignee_id) as assignee,
                    (select name from system.user where user_id=a.supervisor_id) as supervisor,
                    to_char(a.resolved_date,'DD-MM-YYYY HH24:MI:SS') as resolved_date,
                    (select name from system.user where user_id=a.res_dec_by) as resolved_by
                from
                    task.task a
                where
                    project_id=%s
                and status_id=2
            """%data['project_id']).dictresult()

            pending_tasks=db.query("""
                select
                    a.name,
                    to_char(a.assignee_deadline,'DD-MM-YYYY HH24:MI:SS') as assignee_deadline,
                    to_char(a.supervisor_deadline,'DD-MM-YYYY HH24:MI:SS') as supervisor_deadline,
                    to_char(a.deadline,'DD-MM-YYYY HH24:MI:SS') as deadline,
                    (select name from system.user where user_id=a.assignee_id) as assignee,
                    (select name from system.user where user_id=a.supervisor_id) as supervisor
                from
                    task.task a
                where
                    project_id=%s
                and status_id=1
            """%data['project_id']).dictresult()

            inprocess_tasks=db.query("""
                select
                    a.name,
                    to_char(a.assignee_deadline,'DD-MM-YYYY HH24:MI:SS') as assignee_deadline,
                    to_char(a.supervisor_deadline,'DD-MM-YYYY HH24:MI:SS') as supervisor_deadline,
                    to_char(a.deadline,'DD-MM-YYYY HH24:MI:SS') as deadline,
                    (select name from system.user where user_id=a.assignee_id) as assignee,
                    (select name from system.user where user_id=a.supervisor_id) as supervisor,
                    to_char(a.last_updated,'DD-MM-YYYY HH24:MI:SS') as last_updated,
                    (select name from system.user where user_id=a.user_last_updated) as user_last_updated
                from
                    task.task a
                where
                    project_id=%s
                and status_id=6
            """%data['project_id']).dictresult()

            declined_tasks=db.query("""
                select
                    a.name,
                    to_char(a.assignee_deadline,'DD-MM-YYYY HH24:MI:SS') as assignee_deadline,
                    to_char(a.supervisor_deadline,'DD-MM-YYYY HH24:MI:SS') as supervisor_deadline,
                    to_char(a.deadline,'DD-MM-YYYY HH24:MI:SS') as deadline,
                    (select name from system.user where user_id=a.assignee_id) as assignee,
                    (select name from system.user where user_id=a.supervisor_id) as supervisor,
                    (select name from system.user where user_id=a.res_dec_by) as declined_by,
                    to_char(a.last_updated,'DD-MM-YYYY HH24:MI:SS') as declined_date
                from
                    task.task a
                where
                    project_id=%s
                and status_id=3
            """%data['project_id']).dictresult()

            canceled_tasks=db.query("""
                select
                    a.name,
                    to_char(a.assignee_deadline,'DD-MM-YYYY HH24:MI:SS') as assignee_deadline,
                    to_char(a.supervisor_deadline,'DD-MM-YYYY HH24:MI:SS') as supervisor_deadline,
                    to_char(a.deadline,'DD-MM-YYYY HH24:MI:SS') as deadline,
                    (select name from system.user where user_id=a.assignee_id) as assignee,
                    (select name from system.user where user_id=a.supervisor_id) as supervisor,
                    (select name from system.user where user_id=a.user_last_updated) as canceled_by,
                    to_char(a.last_updated,'DD-MM-YYYY HH24:MI:SS') as canceled_date
                from
                    task.task a
                where
                    project_id=%s
                and status_id=5
            """%data['project_id']).dictresult()

            wb = Workbook()
            ws = wb.active
            ftitlebold=Font(name='Arial',size=14,bold=True)
            ftitlebold_align=Alignment(horizontal="right",vertical="center")
            ftitle=Font(name='Arial',size=12)
            ftitlebold12=Font(name='Arial',size=12,bold=True)


            ws.column_dimensions['A'].width = 60
            ws.column_dimensions['B'].width = 23
            ws.column_dimensions['C'].width = 23
            ws.column_dimensions['D'].width = 23
            ws.column_dimensions['E'].width = 23
            ws.column_dimensions['F'].width = 23
            ws.column_dimensions['G'].width = 23
            ws.column_dimensions['H'].width = 23

            title_list=[{'cell':'A1','value':'Proyecto:'},{'cell':'A2','value':'Descripción:'},{'cell':'A3','value':'Creado por:'},{'cell':'A4','value':'Finaliza:'}]
            for x in title_list:
                ws[x['cell']].value=x['value']
                ws[x['cell']].font=ftitlebold
                ws[x['cell']].alignment=ftitlebold_align

            title_value_list=[{'cell':'B1','value':project_info['name']},{'cell':'B2','value':project_info['description']},{'cell':'B3','value':project_info['created_by']},{'cell':'B4','value':project_info['deadline']}]
            for y in title_value_list:
                ws[y['cell']].value=y['value']
                ws[y['cell']].font=ftitle

            ftasktitle=Font(name='Arial',size=12, bold=True, color="ffffff")
            ws.merge_cells('A6:H6')
            ws['A6'].fill=PatternFill("solid",fgColor="0066b3")
            ws['A6'].value='Tareas cerradas'
            ws['A6'].font=ftasktitle
            ws['A6'].alignment=Alignment(horizontal="center",vertical="center")

            row=7
            closed_titles=['Nombre','Auxiliar','Fecha auxiliar','Supervisor','Fecha supervisor','Fecha límite','Resuelta']
            resolved_titles=['Nombre','Auxiliar','Fecha auxiliar','Supervisor','Fecha supervisor','Fecha límite','Resuelta','Resuelta por']
            pending_titles=['Nombre','Auxiliar','Fecha auxiliar','Supervisor','Fecha supervisor','Fecha límite']
            inprocess_titles=['Nombre','Auxiliar','Fecha auxiliar','Supervisor','Fecha supervisor','Fecha límite','Actualizado','Actualizado por']
            declined_titles=['Nombre','Auxiliar','Fecha auxiliar','Supervisor','Fecha supervisor','Fecha límite','Declinada','Declinada por']
            canceled_titles=['Nombre','Auxiliar','Fecha auxiliar','Supervisor','Fecha supervisor','Fecha límite','Cancelada']

            #tareas cerradas
            if closed_tasks!=[]:
                for t in range(1,len(closed_titles)+1):
                    ws.cell(row=row,column=t,value=closed_titles[t-1])
                    ws.cell(row=row,column=t).font=ftitlebold12
                    ws.cell(row=row,column=t).alignment=Alignment(horizontal="center",vertical="center")
                row+=1
                tasks_order=[{'position':1,'value':'name'},{'position':2,'value':'assignee'},{'position':3,'value':'assignee_deadline'},{'position':4,'value':'supervisor'},{'position':5,'value':'supervisor_deadline'},{'position':6,'value':'deadline'},{'position':7,'value':'resolved_date'}]
                for ct in closed_tasks:
                    for to in tasks_order:
                        ws.cell(row=row,column=to['position'],value=ct[to['value']])
                        ws.cell(row=row,column=to['position']).font=ftitle
                        ws.cell(row=row,column=to['position']).alignment=Alignment(horizontal="center",vertical="center")
                    row+=1
                row+=1

            else:
                ws.merge_cells('A7:H7')
                ws['A7'].value='No hay tareas cerradas'
                ws['A7'].alignment=Alignment(horizontal="center",vertical="center")
                ws['A7'].font=ftitle
                row+=2

            merge='A%s:H%s'%(row,row)
            ws.merge_cells(merge)
            ws['A'+str(row)].fill=PatternFill("solid",fgColor="a3238e")
            ws['A'+str(row)].value='Tareas resueltas'
            ws['A'+str(row)].font=ftasktitle
            ws['A'+str(row)].alignment=Alignment(horizontal="center",vertical="center")
            row+=1
            if resolved_tasks!=[]:
                for t in range(1,len(resolved_titles)+1):
                    ws.cell(row=row,column=t,value=resolved_titles[t-1])
                    ws.cell(row=row,column=t).font=ftitlebold12
                    ws.cell(row=row,column=t).alignment=Alignment(horizontal="center",vertical="center")
                row+=1
                tasks_order=[{'position':1,'value':'name'},{'position':2,'value':'assignee'},{'position':3,'value':'assignee_deadline'},{'position':4,'value':'supervisor'},{'position':5,'value':'supervisor_deadline'},{'position':6,'value':'deadline'},{'position':7,'value':'resolved_date'},{'position':8,'value':'resolved_by'}]
                for ct in resolved_tasks:
                    for to in tasks_order:
                        ws.cell(row=row,column=to['position'],value=ct[to['value']])
                        ws.cell(row=row,column=to['position']).font=ftitle
                        ws.cell(row=row,column=to['position']).alignment=Alignment(horizontal="center",vertical="center")
                    row+=1
                row+=1
            else:
                merge='A%s:H%s'%(row,row)
                ws.merge_cells(merge)
                ws['A'+str(row)].value='No hay tareas resueltas'
                ws['A'+str(row)].alignment=Alignment(horizontal="center",vertical="center")
                ws['A'+str(row)].font=ftitle
                row+=2

            merge='A%s:H%s'%(row,row)
            ws.merge_cells(merge)
            ws['A'+str(row)].fill=PatternFill("solid",fgColor="ef413d")
            ws['A'+str(row)].value='Tareas pendientes'
            ws['A'+str(row)].font=ftasktitle
            ws['A'+str(row)].alignment=Alignment(horizontal="center",vertical="center")
            row+=1
            if pending_tasks!=[]:
                for t in range(1,len(pending_titles)+1):
                    ws.cell(row=row,column=t,value=pending_titles[t-1])
                    ws.cell(row=row,column=t).font=ftitlebold12
                    ws.cell(row=row,column=t).alignment=Alignment(horizontal="center",vertical="center")
                row+=1
                tasks_order=[{'position':1,'value':'name'},{'position':2,'value':'assignee'},{'position':3,'value':'assignee_deadline'},{'position':4,'value':'supervisor'},{'position':5,'value':'supervisor_deadline'},{'position':6,'value':'deadline'}]
                for ct in pending_tasks:
                    for to in tasks_order:
                        ws.cell(row=row,column=to['position'],value=ct[to['value']])
                        ws.cell(row=row,column=to['position']).font=ftitle
                        ws.cell(row=row,column=to['position']).alignment=Alignment(horizontal="center",vertical="center")
                    row+=1
                row+=1
            else:
                merge='A%s:H%s'%(row,row)
                ws.merge_cells(merge)
                ws['A'+str(row)].value='No hay tareas pendientes'
                ws['A'+str(row)].alignment=Alignment(horizontal="center",vertical="center")
                ws['A'+str(row)].font=ftitle
                row+=2


            merge='A%s:H%s'%(row,row)
            ws.merge_cells(merge)
            ws['A'+str(row)].fill=PatternFill("solid",fgColor="faa61a")
            ws['A'+str(row)].value='Tareas en proceso'
            ws['A'+str(row)].font=ftasktitle
            ws['A'+str(row)].alignment=Alignment(horizontal="center",vertical="center")
            row+=1
            if inprocess_tasks!=[]:
                for t in range(1,len(inprocess_titles)+1):
                    ws.cell(row=row,column=t,value=inprocess_titles[t-1])
                    ws.cell(row=row,column=t).font=ftitlebold12
                    ws.cell(row=row,column=t).alignment=Alignment(horizontal="center",vertical="center")
                row+=1
                tasks_order=[{'position':1,'value':'name'},{'position':2,'value':'assignee'},{'position':3,'value':'assignee_deadline'},{'position':4,'value':'supervisor'},{'position':5,'value':'supervisor_deadline'},{'position':6,'value':'deadline'},{'position':7,'value':'last_updated'},{'position':8,'value':'user_last_updated'}]
                for ct in inprocess_tasks:
                    for to in tasks_order:
                        ws.cell(row=row,column=to['position'],value=ct[to['value']])
                        ws.cell(row=row,column=to['position']).font=ftitle
                        ws.cell(row=row,column=to['position']).alignment=Alignment(horizontal="center",vertical="center")
                    row+=1
                row+=1
            else:
                merge='A%s:H%s'%(row,row)
                ws.merge_cells(merge)
                ws['A'+str(row)].value='No hay tareas en proceso'
                ws['A'+str(row)].alignment=Alignment(horizontal="center",vertical="center")
                ws['A'+str(row)].font=ftitle
                row+=2

            merge='A%s:H%s'%(row,row)
            ws.merge_cells(merge)
            ws['A'+str(row)].fill=PatternFill("solid",fgColor="72bf44")
            ws['A'+str(row)].value='Tareas declinadas'
            ws['A'+str(row)].font=ftasktitle
            ws['A'+str(row)].alignment=Alignment(horizontal="center",vertical="center")
            row+=1
            if declined_tasks!=[]:
                for t in range(1,len(declined_titles)+1):
                    ws.cell(row=row,column=t,value=declined_titles[t-1])
                    ws.cell(row=row,column=t).font=ftitlebold12
                    ws.cell(row=row,column=t).alignment=Alignment(horizontal="center",vertical="center")
                row+=1
                tasks_order=[{'position':1,'value':'name'},{'position':2,'value':'assignee'},{'position':3,'value':'assignee_deadline'},{'position':4,'value':'supervisor'},{'position':5,'value':'supervisor_deadline'},{'position':6,'value':'deadline'},{'position':7,'value':'declined_date'},{'position':8,'value':'declined_by'}]
                for ct in declined_tasks:
                    for to in tasks_order:
                        ws.cell(row=row,column=to['position'],value=ct[to['value']])
                        ws.cell(row=row,column=to['position']).font=ftitle
                        ws.cell(row=row,column=to['position']).alignment=Alignment(horizontal="center",vertical="center")
                    row+=1
                row+=1
            else:
                merge='A%s:H%s'%(row,row)
                ws.merge_cells(merge)
                ws['A'+str(row)].value='No hay tareas declinadas'
                ws['A'+str(row)].alignment=Alignment(horizontal="center",vertical="center")
                ws['A'+str(row)].font=ftitle
                row+=2

            merge='A%s:H%s'%(row,row)
            ws.merge_cells(merge)
            ws['A'+str(row)].fill=PatternFill("solid",fgColor="684703")
            ws['A'+str(row)].value='Tareas canceladas'
            ws['A'+str(row)].font=ftasktitle
            ws['A'+str(row)].alignment=Alignment(horizontal="center",vertical="center")
            row+=1
            if canceled_tasks!=[]:
                for t in range(1,len(canceled_titles)+1):
                    ws.cell(row=row,column=t,value=canceled_titles[t-1])
                    ws.cell(row=row,column=t).font=ftitlebold12
                    ws.cell(row=row,column=t).alignment=Alignment(horizontal="center",vertical="center")
                row+=1
                tasks_order=[{'position':1,'value':'name'},{'position':2,'value':'assignee'},{'position':3,'value':'assignee_deadline'},{'position':4,'value':'supervisor'},{'position':5,'value':'supervisor_deadline'},{'position':6,'value':'deadline'},{'position':7,'value':'canceled_date'},{'position':8,'value':'canceled_by'}]
                for ct in canceled_tasks:
                    for to in tasks_order:
                        ws.cell(row=row,column=to['position'],value=ct[to['value']])
                        ws.cell(row=row,column=to['position']).font=ftitle
                        ws.cell(row=row,column=to['position']).alignment=Alignment(horizontal="center",vertical="center")
                    row+=1
                row+=1

            else:
                merge='A%s:H%s'%(row,row)
                ws.merge_cells(merge)
                ws['A'+str(row)].value='No hay tareas canceladas'
                ws['A'+str(row)].alignment=Alignment(horizontal="center",vertical="center")
                ws['A'+str(row)].font=ftitle
                row+=2


            fecha=str(datetime.datetime.today())
            fecha=fecha.replace(" ","_")
            fecha=fecha.replace(":","_")
            fecha=fecha.replace(".","_")

            wb.save('%sResumen_%s.xlsx'%(cfg.report_path,fecha))
            response['filename']='/project/downloadProjectSummary/Resumen_%s.xlsx'%fecha
            response['success']=True
            response['msg_response']='El resumen ha sido generado.'

        else:
            response['success']=False
            response['msg_response']='Ocurrió un error al intentar obtener la información del proyecto, favor de intentarlo más tarde.'
    except:
        response['success']=False
        response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo más tarde.'
        exc_info=sys.exc_info()
        app.logger.info(traceback.format_exc(exc_info))
    return json.dumps(response)

@bp.route('/downloadProjectSummary/<filename>', methods=['GET','POST'])
@is_logged_in
def downloadProjectSummary(filename):
    response={}
    try:
        path="%s%s"%(cfg.report_path,filename)
        name="%s"%filename
        return send_file(path,attachment_filename=name)

    except:
        response['success']=False
        response['msg_response']='Ocurrió un error'
        exc_info=sys.exc_info()
        app.logger.info(traceback.format_exc(exc_info))
        return response

@bp.route('/changeProjectStatus', methods=['GET','POST'])
@is_logged_in
def changeProjectStatus():
    response={}
    try:
        flag,data=GF.toDict(request.form,'post')
        if flag:
            resolved=''
            if int(data['status_id'])==2:
                resolved=',resolved_date=now() '

            db.query("""
                update task.project
                set status_id=%s,
                last_updated=now(),
                last_updated_by=%s %s
                where project_id=%s
            """%(data['status_id'],data['user_id'],resolved,data['project_id']))

            response['success']=True
            response['msg_response']='El status ha sido actualizado.'
        else:
            response['success']=False
            response['msg_response']='Ocurrió un error al intentar procesar la información, favor de intentarlo de nuevo.'
    except:
        response['success']=False
        response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo más tarde.'
        exc_info=sys.exc_info()
        app.logger.info(traceback.format_exc(exc_info))
    return json.dumps(response)
