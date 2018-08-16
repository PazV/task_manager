#!/usr/bin/env python
#--*-- coding: utf-8 --*--
from db_connection import getDB
db = getDB()
import generic_functions
GF=generic_functions.GenericFunctions()
import logging
import app_config as cfg
import sys
import traceback
from datetime import date, timedelta, datetime
from dateutil.relativedelta import relativedelta
import smtplib
from email.MIMEMultipart import MIMEMultipart
from email.MIMEText import MIMEText

from os.path import basename
from email.mime.application import MIMEApplication

import re
from openpyxl import Workbook
from openpyxl.styles import colors
from openpyxl.styles import Font, Color


print "Entra do admin report"
# create logger with 'spam_application'
logger = logging.getLogger('Send Notif')
logger.setLevel(logging.DEBUG)
# create file handler which logs even debug messages
fh = logging.FileHandler('%scron_admin_report.log'%cfg.log_path)
fh.setLevel(logging.DEBUG)
# create console handler with a higher log level
ch = logging.StreamHandler()
ch.setLevel(logging.ERROR)
# create formatter and add it to the handlers
formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
fh.setFormatter(formatter)
ch.setFormatter(formatter)
# add the handlers to the logger
if not len(logger.handlers):
    logger.addHandler(fh)
    logger.addHandler(ch)

logger.info("Log info")

class MailFunctions:

    def sendMail(self,to_address,subject,body,file=None):
        success=True
        try:
            server=smtplib.SMTP(cfg.mail_server,cfg.mail_port)
            server.login(cfg.mail_username,cfg.mail_password)
            from_address=cfg.mail_username
            msg=MIMEMultipart()
            msg['From']=from_address
            msg['To']=to_address
            msg['Subject']=subject.decode('utf-8')
            body=self.replaceStringHtml(body)
            msg.attach(MIMEText(body,'html'))

            if file!=None:
                with open(file, "rb") as fil:
                    part = MIMEApplication(
                        fil.read(),
                        Name=basename(file)
                    )
                part['Content-Disposition'] = 'attachment; filename="%s"' % basename(file)
                msg.attach(part)
            text=msg.as_string()
            server.sendmail(from_address,to_address,text)
        except:
            success=False
            exc_info=sys.exc_info()
            logger.error(traceback.format_exc(exc_info))
        return success

    def replaceStringHtml(self,text):
        rep = {
            "á":"&aacute;",
            "é":"&eacute;",
            "í":"&iacute;",
            "ó":"&oacute;",
            "ú":"&uacute;",
            "Á":"&Aacute;",
            "É":"&Eacute;",
            "Í":"&Iacute;",
            "Ó":"&Oacute;",
            "Ú":"&Uacute;",
            "ñ":"&ntilde;",
            "Ñ":"&Ntilde;"
        }
        rep = dict((re.escape(k), v) for k, v in rep.iteritems())
        pattern = re.compile("|".join(rep.keys()))
        new_text = pattern.sub(lambda m: rep[re.escape(m.group(0))], text)
        return new_text

    def as_text(self,value):
        if value is None:
            return ""
        else:
            try:
                return str(value)
            except:
                return value


def main():
    try:
        MF=MailFunctions()
        today=date.today()

        admins=db.query("""
            select
                a.user_id,
                a.name as admin,
                a.company_id,
                a.email,
                (select name from system.company where a.company_id=company_id) as company
            from
                system.user a
            where
                a.user_type_id=1
            and a.enabled in (1,3)
        """).dictresult()

        for a in admins:
            logger.info(a)
            notif_settings=db.query("""
                select
                    admin_report_frequency,
                    to_char(last_admin_notification,'YYYY-MM-DD') as last_admin_notification
                from system.notification_settings
                where company_id=%s
            """%a['company_id']).dictresult()
            do_report=False
            if notif_settings!=[]:
                frequency=notif_settings[0]['admin_report_frequency'].split("_")
                last_admin_notification=notif_settings[0]['last_admin_notification'].split("-")
                d_last_notification=date(int(last_admin_notification[0]),int(last_admin_notification[1]),int(last_admin_notification[2]))
                if frequency[1]=='d': #days
                    if today-timedelta(days=int(frequency[0]))==d_last_notification: #if True, generates a new report
                        do_report=True
                        interval_from=str(today-timedelta(days=int(frequency[0])))
                        interval_to=str(today+timedelta(days=int(frequency[0])))
                elif frequency[1]=='w':
                    if today-timedelta(days=int(frequency[0])*7)==d_last_notification:
                        do_report=True
                        interval_from=str(today-timedelta(days=int(frequency[0])*7))
                        interval_to=str(today+timedelta(days=int(frequency[0])*7))
                elif frequency[1]=='m':
                    if today-relativedelta(months=int(frequency[0]))==d_last_notification:
                        do_report=True
                        interval_from=str(today-relativedelta(months=int(frequency[0])))
                        interval_to=str(today+relativedelta(months=int(frequency[0])))
                else:
                    do_report=False
                if do_report==True:
                    created_tasks=db.query("""
                        select
                            a.name,
                            a.description,
                            to_char(a.deadline,'DD-MM-YYYY') as deadline,
                            (select name from system.user where user_id=a.supervisor_id) as supervisor,
                            to_char(a.supervisor_deadline,'DD-MM-YYYY') as supervisor_deadline,
                            (select name from system.user where user_id=a.assignee_id) as assignee,
                            to_char(a.assignee_deadline,'DD-MM-YYYY') as assignee_deadline,
                            to_char(a.created,'DD-MM-YYYY') as created,
                            (select name from system.user where user_id=a.created_by) as created_by
                        from
                            task.task a
                        where
                            company_id=%s
                        and created between '%s 00:00:00' and '%s 23:59:59'
                    """%(a['company_id'],interval_from,interval_to)).dictresult()
                    created_tasks_title=['Nombre','Descripción','Fecha límite','Supervisor','Fecha límite supervisor','Auxiliar','Fecha límite auxiliar','Fecha de creación','Creado por']
                    created_tasks_d_order=['name','description','deadline','supervisor','supervisor_deadline','assignee','assignee_deadline','created','created_by']

                    expired_assignee_tasks=db.query("""
                        select
                            a.name,
                            a.description,
                            to_char(a.deadline,'DD-MM-YYYY') as deadline,
                            (select name from system.user where user_id=a.supervisor_id) as supervisor,
                            to_char(a.supervisor_deadline,'DD-MM-YYYY') as supervisor_deadline,
                            (select name from system.user where user_id=a.assignee_id) as assignee,
                            to_char(a.assignee_deadline,'DD-MM-YYYY') as assignee_deadline,
                            to_char(a.created,'DD-MM-YYYY') as created,
                            (select name from system.user where user_id=a.created_by) as created_by,
                            to_char(a.last_updated, 'DD-MM-YYYY') as last_updated,
                            (select name from system.user where user_id=a.user_last_updated) as user_last_updated
                        from
                            task.task a
                        where
                            company_id=%s
                        and assignee_deadline between '%s 00:00:00' and '%s 23:59:59'
                        and status_id in (1,6)
                    """%(a['company_id'],interval_from,interval_to)).dictresult()

                    expired_assignee_tasks_title=['Nombre','Descripción','Fecha límite','Supervisor','Fecha límite supervisor','Auxiliar','Fecha límite auxiliar','Fecha de creación','Creado por','Actualizado por última vez','Actualizado por']
                    expired_assignee_tasks_d_order=['name','description','deadline','supervisor','supervisor_deadline','assignee','assignee_deadline','created','created_by','last_updated','user_last_updated']

                    expired_supervisor_tasks=db.query("""
                        select
                            a.name,
                            a.description,
                            to_char(a.deadline,'DD-MM-YYYY') as deadline,
                            (select name from system.user where user_id=a.supervisor_id) as supervisor,
                            to_char(a.supervisor_deadline,'DD-MM-YYYY') as supervisor_deadline,
                            (select name from system.user where user_id=a.assignee_id) as assignee,
                            to_char(a.assignee_deadline,'DD-MM-YYYY') as assignee_deadline,
                            to_char(a.created,'DD-MM-YYYY') as created,
                            (select name from system.user where user_id=a.created_by) as created_by,
                            to_char(a.last_updated, 'DD-MM-YYYY') as last_updated,
                            (select name from system.user where user_id=a.user_last_updated) as user_last_updated
                        from
                            task.task a
                        where
                            company_id=%s
                        and supervisor_deadline between '%s 00:00:00' and '%s 23:59:59'
                        and status_id in (1,6)
                    """%(a['company_id'],interval_from,interval_to)).dictresult()

                    expired_supervisor_tasks_title=['Nombre','Descripción','Fecha límite','Supervisor','Fecha límite supervisor','Auxiliar','Fecha límite auxiliar','Fecha de creación','Creado por','Actualizado por última vez','Actualizado por']
                    expired_supervisor_tasks_d_order=['name','description','deadline','supervisor','supervisor_deadline','assignee','assignee_deadline','created','created_by','last_updated','user_last_updated']

                    expired_admin_tasks=db.query("""
                        select
                            a.name,
                            a.description,
                            to_char(a.deadline,'DD-MM-YYYY') as deadline,
                            (select name from system.user where user_id=a.supervisor_id) as supervisor,
                            to_char(a.supervisor_deadline,'DD-MM-YYYY') as supervisor_deadline,
                            (select name from system.user where user_id=a.assignee_id) as assignee,
                            to_char(a.assignee_deadline,'DD-MM-YYYY') as assignee_deadline,
                            to_char(a.created,'DD-MM-YYYY') as created,
                            (select name from system.user where user_id=a.created_by) as created_by,
                            to_char(a.last_updated, 'DD-MM-YYYY') as last_updated,
                            (select name from system.user where user_id=a.user_last_updated) as user_last_updated
                        from
                            task.task a
                        where
                            company_id=%s
                        and deadline between '%s 00:00:00' and '%s 23:59:59'
                        and status_id in (1,6)
                    """%(a['company_id'],interval_from,interval_to)).dictresult()

                    expired_admin_tasks_title=['Nombre','Descripción','Fecha límite','Supervisor','Fecha límite supervisor','Auxiliar','Fecha límite auxiliar','Fecha de creación','Creado por','Actualizado por última vez','Actualizado por']
                    expired_admin_tasks_d_order=['name','description','deadline','supervisor','supervisor_deadline','assignee','assignee_deadline','created','created_by','last_updated','user_last_updated']

                    resolved_tasks=db.query("""
                        select
                            a.name,
                            a.description,
                            to_char(a.deadline,'DD-MM-YYYY') as deadline,
                            (select name from system.user where user_id=a.supervisor_id) as supervisor,
                            to_char(a.supervisor_deadline,'DD-MM-YYYY') as supervisor_deadline,
                            (select name from system.user where user_id=a.assignee_id) as assignee,
                            to_char(a.assignee_deadline,'DD-MM-YYYY') as assignee_deadline,
                            to_char(a.created,'DD-MM-YYYY') as created,
                            (select name from system.user where user_id=a.created_by) as created_by,
                            to_char(a.last_updated, 'DD-MM-YYYY') as last_updated,
                            (select name from system.user where user_id=a.user_last_updated) as user_last_updated,
                            to_char(a.resolved_date,'DD-MM-YYYY') as resolved_date,
                            a.comments
                        from
                            task.task a
                        where
                            company_id=%s
                        and resolved_date between '%s 00:00:00' and '%s 23:59:59'
                        and status_id = 2
                    """%(a['company_id'],interval_from,interval_to)).dictresult()

                    resolved_tasks_title=['Nombre','Descripción','Fecha límite','Supervisor','Fecha límite supervisor','Auxiliar','Fecha límite auxiliar','Fecha de creación','Creado por','Actualizado por última vez','Actualizado por','Fecha en que se resolvió','Comentarios']
                    resolved_tasks_d_order=['name','description','deadline','supervisor','supervisor_deadline','assignee','assignee_deadline','created','created_by','last_updated','user_last_updated','resolved_date','comments']

                    closed_tasks=db.query("""
                        select
                            a.name,
                            a.description,
                            to_char(a.deadline,'DD-MM-YYYY') as deadline,
                            (select name from system.user where user_id=a.supervisor_id) as supervisor,
                            to_char(a.supervisor_deadline,'DD-MM-YYYY') as supervisor_deadline,
                            (select name from system.user where user_id=a.assignee_id) as assignee,
                            to_char(a.assignee_deadline,'DD-MM-YYYY') as assignee_deadline,
                            to_char(a.created,'DD-MM-YYYY') as created,
                            (select name from system.user where user_id=a.created_by) as created_by,
                            to_char(a.last_updated, 'DD-MM-YYYY') as last_updated,
                            (select name from system.user where user_id=a.user_last_updated) as user_last_updated,
                            to_char(a.resolved_date,'DD-MM-YYYY') as resolved_date,
                            a.comments,
                            a.supervisor_comments
                        from
                            task.task a
                        where
                            company_id=%s
                        and last_updated between '%s 00:00:00' and '%s 23:59:59'
                        and status_id = 4
                    """%(a['company_id'],interval_from,interval_to)).dictresult()

                    closed_tasks_title=['Nombre','Descripción','Fecha límite','Supervisor','Fecha límite supervisor','Auxiliar','Fecha límite auxiliar','Fecha de creación','Creado por','Actualizado por última vez','Actualizado por','Fecha en que se resolvió','Comentarios','Comentarios del supervisor']
                    closed_tasks_d_order=['name','description','deadline','supervisor','supervisor_deadline','assignee','assignee_deadline','created','created_by','last_updated','user_last_updated','resolved_date','comments','supervisor_comments']

                    declined_tasks=db.query("""
                        select
                            a.name,
                            a.description,
                            to_char(a.deadline,'DD-MM-YYYY') as deadline,
                            (select name from system.user where user_id=a.supervisor_id) as supervisor,
                            to_char(a.supervisor_deadline,'DD-MM-YYYY') as supervisor_deadline,
                            (select name from system.user where user_id=a.assignee_id) as assignee,
                            to_char(a.assignee_deadline,'DD-MM-YYYY') as assignee_deadline,
                            to_char(a.created,'DD-MM-YYYY') as created,
                            (select name from system.user where user_id=a.created_by) as created_by,
                            to_char(a.last_updated, 'DD-MM-YYYY') as last_updated,
                            (select name from system.user where user_id=a.user_last_updated) as user_last_updated,
                            a.declining_cause
                        from
                            task.task a
                        where
                            company_id=%s
                        and last_updated between '%s 00:00:00' and '%s 23:59:59'
                        and status_id = 3
                    """%(a['company_id'],interval_from,interval_to)).dictresult()

                    declined_tasks_title=['Nombre','Descripción','Fecha límite','Supervisor','Fecha límite supervisor','Auxiliar','Fecha límite auxiliar','Fecha de creación','Creado por','Actualizado por última vez','Actualizado por','Motivo para declinar']
                    declined_tasks_d_order=['name','description','deadline','supervisor','supervisor_deadline','assignee','assignee_deadline','created','created_by','last_updated','user_last_updated','declining_cause']

                    canceled_tasks=db.query("""
                        select
                            a.name,
                            a.description,
                            to_char(a.deadline,'DD-MM-YYYY') as deadline,
                            (select name from system.user where user_id=a.supervisor_id) as supervisor,
                            to_char(a.supervisor_deadline,'DD-MM-YYYY') as supervisor_deadline,
                            (select name from system.user where user_id=a.assignee_id) as assignee,
                            to_char(a.assignee_deadline,'DD-MM-YYYY') as assignee_deadline,
                            to_char(a.created,'DD-MM-YYYY') as created,
                            (select name from system.user where user_id=a.created_by) as created_by,
                            to_char(a.last_updated, 'DD-MM-YYYY') as last_updated,
                            (select name from system.user where user_id=a.user_last_updated) as user_last_updated,
                            a.declining_cause
                        from
                            task.task a
                        where
                            company_id=%s
                        and last_updated between '%s 00:00:00' and '%s 23:59:59'
                        and status_id = 5
                    """%(a['company_id'],interval_from,interval_to)).dictresult()

                    canceled_tasks_title=['Nombre','Descripción','Fecha límite','Supervisor','Fecha límite supervisor','Auxiliar','Fecha límite auxiliar','Fecha de creación','Creado por','Actualizado por última vez','Actualizado por','Motivo para declinar']
                    canceled_tasks_d_order=['name','description','deadline','supervisor','supervisor_deadline','assignee','assignee_deadline','created','created_by','last_updated','user_last_updated','declining_cause']

                    list_list=[
                        {'title':created_tasks_title,'data':created_tasks,'sheet_name':'Tareas creadas','data_order':created_tasks_d_order,'color':"000040ff"},
                        {'title':expired_assignee_tasks_title, 'data':expired_assignee_tasks,'sheet_name':'Tareas aux exp','data_order':expired_assignee_tasks_d_order,'color':"00ffd11a"},
                        {'title':expired_supervisor_tasks_title, 'data':expired_supervisor_tasks, 'sheet_name':'Tareas sup exp','data_order':expired_supervisor_tasks_d_order,'color':"00ff8000"},
                        {'title':expired_admin_tasks_title, 'data':expired_admin_tasks,'sheet_name':'Tareas exp','data_order':expired_admin_tasks_d_order,'color':"00e62e00"},
                        {'title':resolved_tasks_title, 'data':resolved_tasks,'sheet_name':'Resueltas','data_order':resolved_tasks_d_order,'color':"0000e6b8"},
                        {'title':closed_tasks_title, 'data':closed_tasks,'sheet_name':'Cerradas','data_order':closed_tasks_d_order,'color':"0000b300"},
                        {'title':declined_tasks_title, 'data':declined_tasks,'sheet_name':'Declinadas','data_order':declined_tasks_d_order,'color':"00cc00cc"},
                        {'title':canceled_tasks_title, 'data':canceled_tasks,'sheet_name':'Canceladas','data_order':canceled_tasks_d_order,'color':"00999999"}
                    ]

                    wb = Workbook()
                    sheet_number = 0
                    for ll in list_list:
                        logger.info("------------Haciendo %s-----------------"%ll['sheet_name'])
                        if ll['data']!=[]:
                            ws = wb.create_sheet("%s"%ll['sheet_name'],sheet_number)
                            ws.sheet_properties.tabColor = ll['color']
                            sheet_number+=1
                            for i in range(1,len(ll['title'])+1):
                                ws.cell(row=1,column=i,value=ll['title'][i-1])
                                ws.cell(row=1,column=i).font=Font(name='Arial', size=12, bold=True)
                            row=2
                            for x in ll['data']:
                                logger.info(x)
                                for r in range(0,len(ll['data_order'])):
                                    logger.info("%s : %s"%(ll['data_order'][r],x[ll['data_order'][r]]))
                                    ws.cell(row=row,column=r+1,value=x[ll['data_order'][r]])
                                    ws.cell(row=row,column=r+1).font=Font(name='Arial',size=11)
                                row+=1

                    for w in wb.sheetnames:
                        sheet=wb[w]
                        for column_cells in sheet.columns:
                            length = max(len(MF.as_text(cell.value))+5 for cell in column_cells)
                            sheet.column_dimensions[column_cells[0].column].width = length
                    fecha=str(datetime.today())
                    fecha=fecha.replace(" ","_")
                    fecha=fecha.replace(":","_")
                    fecha=fecha.replace(".","_")

                    new_from=interval_from.split("-")
                    new_to=interval_to.split("-")
                    mail_info={
                        'admin':a['admin'],
                        'company':a['company'],
                        'from':'%s-%s-%s'%(new_from[2],new_from[1],new_from[0]),
                        'to':'%s-%s-%s'%(new_to[2],new_to[1],new_to[0]),
                        'link':cfg.host
                    }
                    if len(wb.sheetnames)>1:
                        wb.remove(wb['Sheet'])
                        wb.save('/tmp/Reporte_%s.xlsx'%fecha)
                        mail_template=db.query("""
                            select * from template.generic_template where type_id=22
                        """).dictresult()[0]
                        msg_body=mail_template['body'].format(**mail_info)
                        MF.sendMail(a['email'],mail_template['subject'],msg_body,'/tmp/Reporte_%s.xlsx'%fecha)
                        logger.info("Genera reporte de empresa %s"%a['company'])
                    else:
                        mail_template=db.query("""
                            select * from template.generic_template where type_id=23
                        """).dictresult()[0]
                        msg_body=mail_template['body'].format(**mail_info)
                        MF.sendMail(a['email'],mail_template['subject'],msg_body)
                        logger.info("Reporte vacío de empresa %s"%a['company'])

                    db.query("""
                        update system.notification_settings set last_admin_notification = now() where company_id=%s
                    """%a['company_id'])

            else:
                logger.info("La empresa company_id %s no tiene configuración de notificaciones."%a['company_id'])

        print "Termina reporte administrador"

    except:
        exc_info = sys.exc_info()
        logger.error(traceback.format_exc(exc_info))

if __name__ == '__main__':
    main()
